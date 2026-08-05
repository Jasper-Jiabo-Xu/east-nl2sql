#!/usr/bin/env python3
"""Build review-only, source-traceable V2 code-reference and local-value SQLite libraries.

This script deliberately does not consume Agent1's V2 draft output.  The draft has a
known `CODE_DOMAIN` granularity defect and must not become a source of code values.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from vnext.constraint_assets_v1.compile_tasks import ATT3, parse_att3


ROOT = Path(__file__).resolve().parents[2]
WORKING = ROOT / "kb/working/20260727_000_goal_mode_v1"
DEFAULT_OUTPUT = WORKING / "构建过程层/04_字段多字段与对象明细状态提取-V2/code_value_libraries/20260804_pre_agent1_v3"
CODEBOOK = ROOT / "data/raw/20260626_east_materials/east材料/规范附件4：金融资产管理公司监管数据标准化规范业务代码表.xlsx"
REVIEW_BOOK = WORKING / "构建过程层/01_附件3主数据与附件2技术规则解析/附件3附件2解析审阅报告.xlsx"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha(value: Any) -> str:
    return hashlib.sha256(canonical_json(value).encode("utf-8")).hexdigest()


def ref(source: str, sheet: str, row: int, last_column: str) -> str:
    return f"{source}:{sheet}!A{row}:{last_column}{row}"


def cell_ref(source: str, sheet: str, row: int, column: str) -> str:
    return f"{source}:{sheet}!{column}{row}"


def row_from_row_ref(source_ref: str) -> int:
    matched = re.search(r"![A-Z]+(\d+):[A-Z]+\1$", source_ref)
    if not matched:
        raise RuntimeError(f"无法解析来源行号: {source_ref}")
    return int(matched.group(1))


def normalize_standard(value: str) -> str:
    cleaned = re.sub(r"\s+", "", value.upper())
    cleaned = cleaned.replace("（", "(").replace("）", ")")
    return cleaned


def make_conn(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=DELETE")
    return conn


def create_reference_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE source_document (
          source_id TEXT PRIMARY KEY,
          absolute_path TEXT NOT NULL,
          sha256 TEXT NOT NULL,
          imported_at TEXT NOT NULL
        );
        CREATE TABLE external_standard (
          external_standard_id TEXT PRIMARY KEY,
          east_citation TEXT NOT NULL,
          standard_name TEXT NOT NULL,
          verification_status TEXT NOT NULL,
          source_summary TEXT NOT NULL,
          authoritative_link TEXT,
          use_decision TEXT NOT NULL,
          source_ref TEXT NOT NULL UNIQUE
        );
        CREATE TABLE code_table (
          code_table_id TEXT PRIMARY KEY,
          table_name TEXT NOT NULL,
          sheet_name TEXT NOT NULL UNIQUE,
          reference_standard_raw TEXT NOT NULL,
          authority_key TEXT NOT NULL,
          directory_source_ref TEXT NOT NULL UNIQUE,
          header_row INTEGER,
          code_headers_json TEXT NOT NULL,
          name_header TEXT,
          source_id TEXT NOT NULL REFERENCES source_document(source_id)
        );
        CREATE TABLE code_table_source_row (
          code_table_id TEXT NOT NULL REFERENCES code_table(code_table_id),
          source_row INTEGER NOT NULL,
          source_ref TEXT NOT NULL UNIQUE,
          raw_row_json TEXT NOT NULL,
          PRIMARY KEY (code_table_id, source_row)
        );
        CREATE TABLE canonical_code_value (
          code_value_id TEXT PRIMARY KEY,
          authority_key TEXT NOT NULL,
          code_components_json TEXT NOT NULL,
          display_name TEXT NOT NULL,
          description TEXT NOT NULL,
          value_fingerprint TEXT NOT NULL UNIQUE
        );
        CREATE TABLE code_table_value (
          code_table_id TEXT NOT NULL REFERENCES code_table(code_table_id),
          code_value_id TEXT NOT NULL REFERENCES canonical_code_value(code_value_id),
          source_row INTEGER NOT NULL,
          source_ref TEXT NOT NULL UNIQUE,
          PRIMARY KEY (code_table_id, code_value_id, source_row)
        );
        CREATE TABLE code_value_alias (
          code_table_id TEXT NOT NULL REFERENCES code_table(code_table_id),
          code_value_id TEXT NOT NULL REFERENCES canonical_code_value(code_value_id),
          alias_kind TEXT NOT NULL,
          alias_value TEXT NOT NULL,
          PRIMARY KEY (code_table_id, alias_kind, alias_value)
        );
        CREATE TABLE data_element_code_table_binding (
          binding_id TEXT PRIMARY KEY,
          external_standard_id TEXT NOT NULL REFERENCES external_standard(external_standard_id),
          east_citation TEXT NOT NULL,
          data_element_code TEXT NOT NULL,
          data_element_name TEXT NOT NULL,
          att3_location TEXT NOT NULL,
          att3_quote TEXT NOT NULL,
          code_table_id TEXT REFERENCES code_table(code_table_id),
          binding_status TEXT NOT NULL CHECK(binding_status IN ('BOUND','NO_ATT4_TABLE')),
          source_ref TEXT NOT NULL UNIQUE,
          UNIQUE(external_standard_id, data_element_code, code_table_id)
        );
        CREATE INDEX idx_code_value_authority ON canonical_code_value(authority_key);
        CREATE INDEX idx_code_alias_lookup ON code_value_alias(code_table_id, alias_value);
        CREATE INDEX idx_binding_data_element ON data_element_code_table_binding(data_element_code);
        """
    )


def create_local_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE source_document (
          source_id TEXT PRIMARY KEY,
          absolute_path TEXT NOT NULL,
          sha256 TEXT NOT NULL,
          imported_at TEXT NOT NULL
        );
        CREATE TABLE local_value_source (
          local_source_id TEXT PRIMARY KEY,
          owner_kind TEXT NOT NULL CHECK(owner_kind IN ('DATA_ELEMENT_DESCRIPTION','FIELD_REMARK')),
          owner_id TEXT NOT NULL,
          owner_name TEXT NOT NULL,
          source_ref TEXT NOT NULL UNIQUE,
          raw_text TEXT NOT NULL,
          extraction_status TEXT NOT NULL CHECK(extraction_status IN ('PENDING_AGENT1_V3','NO_LOCAL_VALUE_SET')),
          source_id TEXT NOT NULL REFERENCES source_document(source_id),
          UNIQUE(owner_kind, owner_id)
        );
        CREATE TABLE local_code_set (
          local_code_set_id TEXT PRIMARY KEY,
          owner_kind TEXT NOT NULL,
          owner_id TEXT NOT NULL,
          set_name TEXT NOT NULL,
          source_ref TEXT NOT NULL,
          extraction_contract_version TEXT NOT NULL,
          extraction_status TEXT NOT NULL CHECK(extraction_status IN ('PENDING','EXTRACTED','MANUAL_REVIEW')),
          UNIQUE(owner_kind, owner_id, set_name)
        );
        CREATE TABLE local_code_value (
          local_code_value_id TEXT PRIMARY KEY,
          local_code_set_id TEXT NOT NULL REFERENCES local_code_set(local_code_set_id),
          literal_value TEXT NOT NULL,
          value_code TEXT NOT NULL DEFAULT '',
          value_label TEXT NOT NULL DEFAULT '',
          condition_json TEXT NOT NULL,
          source_ref TEXT NOT NULL,
          value_fingerprint TEXT NOT NULL UNIQUE,
          UNIQUE(local_code_set_id, literal_value, value_code, value_label)
        );
        """
    )


def header_index(values: list[str]) -> tuple[int | None, list[int], int | None, list[int]]:
    for row_no, row in enumerate(values[:10], start=1):
        cells = [norm(value) for value in row]
        code_cols = [index for index, cell in enumerate(cells) if "代码" in cell or cell == "编码"]
        name_cols = [index for index, cell in enumerate(cells) if "名称" in cell or cell == "名"]
        if code_cols and name_cols:
            description_cols = [index for index, cell in enumerate(cells) if "说明" in cell or "备注" in cell]
            return row_no, code_cols, name_cols[0], description_cols
    return None, [], None, []


def build_reference_library(path: Path) -> dict[str, Any]:
    conn = make_conn(path)
    create_reference_schema(conn)
    imported_at = now()
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_AMC_CODE", str(CODEBOOK.resolve()), hash_file(CODEBOOK), imported_at))
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_REVIEW", str(REVIEW_BOOK.resolve()), hash_file(REVIEW_BOOK), imported_at))

    review = load_workbook(REVIEW_BOOK, read_only=True, data_only=True)
    standard_sheet = review["外部标准候选"]
    headers = [norm(cell) for cell in next(standard_sheet.iter_rows(values_only=True))]
    for row_no, row in enumerate(standard_sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [norm(value) for value in row[:7]]
        if not values[0]:
            continue
        conn.execute("INSERT INTO external_standard VALUES (?,?,?,?,?,?,?,?)", (*values, ref("SRC_REVIEW", "外部标准候选", row_no, "G")))

    codebook = load_workbook(CODEBOOK, read_only=True, data_only=True)
    directory = codebook["目录"]
    directory_map: dict[str, tuple[str, str, int]] = {}
    for row_no, row in enumerate(directory.iter_rows(min_row=3, values_only=True), start=3):
        code, name, standard = (norm(row[0]), norm(row[1]), norm(row[2]))
        if code and name:
            directory_map[name] = (code, standard, row_no)

    table_count = source_row_count = value_count = alias_count = 0
    for sheet_name in codebook.sheetnames:
        if sheet_name == "目录":
            continue
        code_table_id, standard, directory_row = directory_map.get(sheet_name, (f"SHEET_{sha(sheet_name)[:12]}", "未在目录列示", 0))
        rows = [[norm(value) for value in row] for row in codebook[sheet_name].iter_rows(values_only=True)]
        header_row, code_columns, name_column, description_columns = header_index(rows)
        headers_row = rows[header_row - 1] if header_row else []
        code_headers = [headers_row[index] for index in code_columns]
        name_header = headers_row[name_column] if name_column is not None else None
        authority_key = normalize_standard(standard) if standard and "自定义" not in standard else f"AMC_CODE:{code_table_id}"
        directory_ref = ref("SRC_AMC_CODE", "目录", directory_row, "C") if directory_row else f"SRC_AMC_CODE:{sheet_name}!A1"
        conn.execute("INSERT INTO code_table VALUES (?,?,?,?,?,?,?,?,?,?)", (code_table_id, sheet_name, sheet_name, standard, authority_key, directory_ref, header_row, canonical_json(code_headers), name_header, "SRC_AMC_CODE"))
        table_count += 1
        for row_no, row in enumerate(rows, start=1):
            if not any(row):
                continue
            source_ref = ref("SRC_AMC_CODE", sheet_name, row_no, chr(64 + max(1, len(row))))
            conn.execute("INSERT INTO code_table_source_row VALUES (?,?,?,?)", (code_table_id, row_no, source_ref, canonical_json(row)))
            source_row_count += 1
            if not header_row or row_no <= header_row:
                continue
            code_components = {headers_row[index]: row[index] for index in code_columns if index < len(row) and row[index]}
            display_name = row[name_column] if name_column is not None and name_column < len(row) else ""
            if not code_components or not display_name or display_name == "返回目录":
                continue
            description = "；".join(row[index] for index in description_columns if index < len(row) and row[index])
            fingerprint = sha({"authority": authority_key, "codes": code_components, "name": display_name, "description": description})
            code_value_id = "CV_" + fingerprint[:24]
            conn.execute("INSERT OR IGNORE INTO canonical_code_value VALUES (?,?,?,?,?,?)", (code_value_id, authority_key, canonical_json(code_components), display_name, description, fingerprint))
            conn.execute("INSERT OR IGNORE INTO code_table_value VALUES (?,?,?,?)", (code_table_id, code_value_id, row_no, source_ref))
            value_count += int(conn.execute("SELECT changes()").fetchone()[0])
            for header, alias in code_components.items():
                conn.execute("INSERT OR IGNORE INTO code_value_alias VALUES (?,?,?,?)", (code_table_id, code_value_id, header, alias))
                alias_count += int(conn.execute("SELECT changes()").fetchone()[0])

    binding_sheet = review["国标与数据元绑定"]
    for row_no, row in enumerate(binding_sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [norm(value) for value in row[:7]]
        if not values[0]:
            continue
        external_id, citation, element_code, element_name, location, quote, code_table_id = values
        status = "BOUND" if code_table_id else "NO_ATT4_TABLE"
        binding_id = "BND_" + sha({"external": external_id, "element": element_code, "table": code_table_id})[:24]
        conn.execute("INSERT INTO data_element_code_table_binding VALUES (?,?,?,?,?,?,?,?,?,?)", (binding_id, external_id, citation, element_code, element_name, location, quote, code_table_id or None, status, ref("SRC_REVIEW", "国标与数据元绑定", row_no, "G")))

    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
    duplicate_aliases = conn.execute("SELECT COUNT(*) FROM (SELECT code_table_id, alias_kind, alias_value, COUNT(*) n FROM code_value_alias GROUP BY 1,2,3 HAVING n > 1)").fetchone()[0]
    counts = {table: conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] for table in ("external_standard", "code_table", "code_table_source_row", "canonical_code_value", "code_table_value", "code_value_alias", "data_element_code_table_binding")}
    conn.commit()
    conn.close()
    return {"integrity_check": integrity, "duplicate_alias_conflicts": duplicate_aliases, "attempted_value_rows": value_count, "attempted_aliases": alias_count, **counts}


def build_local_library(path: Path) -> dict[str, Any]:
    conn = make_conn(path)
    create_local_schema(conn)
    imported_at = now()
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_ATT3", str(ATT3.resolve()), hash_file(ATT3), imported_at))
    elements, _, _ = parse_att3()
    for element in elements.values():
        row = row_from_row_ref(element["source_refs"][0])
        source_ref = cell_ref("SRC_ATT3", "数据元说明", row, "E")
        source_id = "LVS_" + sha({"kind": "DATA_ELEMENT_DESCRIPTION", "code": element["code"]})[:24]
        conn.execute("INSERT INTO local_value_source VALUES (?,?,?,?,?,?,?,?)", (source_id, "DATA_ELEMENT_DESCRIPTION", element["code"], element["name"], source_ref, element["description"], "PENDING_AGENT1_V3", "SRC_ATT3"))
    workbook = load_workbook(ATT3, read_only=True, data_only=True)
    active_table_code = ""
    for row_no, row in enumerate(workbook["数据结构"].iter_rows(values_only=True), start=1):
        values = list(row[:16])
        table_code = norm(values[2] if len(values) > 2 else "")
        if re.fullmatch(r"[A-Z][A-Z0-9]*", table_code):
            active_table_code = table_code
        field_code = norm(values[6] if len(values) > 6 else "")
        field_name = norm(values[5] if len(values) > 5 else "")
        remarks = norm(values[12] if len(values) > 12 else "")
        if not active_table_code or not re.fullmatch(r"[A-Z][A-Z0-9]*", field_code) or not remarks:
            continue
        field_id = f"FLD_{active_table_code}_{field_code}"
        source_ref = cell_ref("SRC_ATT3", "数据结构", row_no, "M")
        source_id = "LVS_" + sha({"kind": "FIELD_REMARK", "id": field_id})[:24]
        conn.execute("INSERT INTO local_value_source VALUES (?,?,?,?,?,?,?,?)", (source_id, "FIELD_REMARK", field_id, field_name, source_ref, remarks, "PENDING_AGENT1_V3", "SRC_ATT3"))
    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
    count = conn.execute("SELECT COUNT(*) FROM local_value_source").fetchone()[0]
    conn.commit()
    conn.close()
    return {"integrity_check": integrity, "local_value_source_count": count, "local_code_set_count": 0, "local_code_value_count": 0}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    reference_db = output / "code_reference_library_v2.sqlite"
    local_db = output / "local_code_value_library_v2.sqlite"
    if reference_db.exists() or local_db.exists():
        raise SystemExit("目标SQLite已存在；为避免覆盖审查资产，请指定新的--output-dir")
    audit = {
        "created_at": now(),
        "publication_status": "review_only_not_released",
        "sources": {
            "attachment4_codebook": str(CODEBOOK.relative_to(ROOT)),
            "existing_standard_binding_review": str(REVIEW_BOOK.relative_to(ROOT)),
        },
        "reference_library": build_reference_library(reference_db),
        "local_value_library": build_local_library(local_db),
        "next_step": "经人工确认后，V3 Agent1 才可向local_code_set/local_code_value写入原子枚举和编码分段结果。",
    }
    (output / "code_value_library_audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
