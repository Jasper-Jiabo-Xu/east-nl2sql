#!/usr/bin/env python3
"""Create extension V2 by adding approved chart-of-accounts bindings to V1.

The V1 extension remains immutable.  V2 is a byte-copy of V1 plus only three
new `tree_reference_usage` records for EAST data elements 003002--003004 and
new extension metadata/audit information.
"""

from __future__ import annotations

import hashlib
import json
import shutil
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[2]
PHASE = ROOT / "kb/working/20260727_000_goal_mode_v1/构建过程层/04_字段多字段与对象明细状态提取-V2"
V1_DB = PHASE / "extension_assets_v1/20260805_扩展数据元约束资产V1_机构部门树绑定/constraint_asset_extension_v1.sqlite"
ACCOUNT_TREE_DB = PHASE / "extension_assets_v1/20260805_会计科目树V1_人工审批后/synthetic_chart_of_accounts_tree_v1.sqlite"
OUTPUT_DIR = PHASE / "extension_assets_v1/20260805_扩展数据元约束资产V2_机构部门与会计科目树绑定"
OUTPUT_DB = OUTPUT_DIR / "constraint_asset_extension_v2.sqlite"

ASSET_VERSION = "constraint-assets-extension-v2-org-department-chart-20260805"
TREE_ASSET_ID = "SYNTHETIC_CHART_OF_ACCOUNTS_TREE_V1"
DESIGN_REF = "SRC_EXTENSION_DESIGN:扩展数据元约束资产V2_会计科目树绑定"
APPROVAL_REF = "SRC_HUMAN_APPROVAL:20260805_科目树四级扩展合同及发布批准"


def canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def assert_ready() -> dict[str, str]:
    if not V1_DB.exists():
        raise RuntimeError(f"缺少扩展V1：{V1_DB}")
    if not ACCOUNT_TREE_DB.exists():
        raise RuntimeError(f"缺少已审批科目树：{ACCOUNT_TREE_DB}")
    tree = sqlite3.connect(ACCOUNT_TREE_DB)
    try:
        meta = dict(tree.execute("SELECT asset_key,asset_value FROM asset_meta").fetchall())
        if meta.get("asset_status") != "approved_not_released":
            raise RuntimeError("科目树未获审批")
        if tree.execute("SELECT COUNT(*) FROM account_node WHERE review_status<>'APPROVED'").fetchone()[0]:
            raise RuntimeError("科目树仍有未审批节点")
        if tree.execute("SELECT COUNT(*) FROM east_data_element_binding_plan WHERE review_status<>'APPROVED' OR dependency_status<>'READY'").fetchone()[0]:
            raise RuntimeError("科目树数据元绑定计划尚未就绪")
        return meta
    finally:
        tree.close()


def refs_and_quote(conn: sqlite3.Connection, code: str) -> tuple[str, str]:
    row = conn.execute("SELECT description,format_source_refs_json FROM data_element WHERE data_element_code=?", (code,)).fetchone()
    if not row:
        raise RuntimeError(f"缺少数据元：{code}")
    return row[0], canonical(json.loads(row[1]) + [DESIGN_REF, APPROVAL_REF])


def record(ref_id: str, code: str, projection: list[str], notes: str) -> dict[str, Any]:
    return {
        "tree_reference_id": ref_id,
        "data_element_code": code,
        "tree_asset_id": TREE_ASSET_ID,
        "tree_sqlite_path_relative": str(ACCOUNT_TREE_DB.relative_to(PHASE)),
        "tree_sqlite_sha256": sha256_file(ACCOUNT_TREE_DB),
        "target_table": "account_node",
        "node_filter_json": canonical({"status_code": "ACTIVE", "is_postable": 1, "posting_policy": ["POSTABLE", "POSTABLE_AND_EXPANDABLE"]}),
        "lookup_key_column": "account_code",
        "value_projection_json": canonical(projection),
        "representation_template": "{" + projection[0] + "}",
        "value_translation_json": canonical({}),
        "binding_mode": "DIRECT_VALUE",
        "applicability_condition_json": canonical({"record_uses_general_ledger_account": True}),
        "dependency_status": "READY",
        "fallback_policy": "NO_FALLBACK_MANUAL_REVIEW",
        "evidence_quote": "",
        "source_refs_json": "",
        "provenance_type": "EXTENSION_DESIGN",
        "review_status": "APPROVED",
        "notes": notes,
    }


RECORDS = (
    record("TREE_REF_003002", "003002", ["account_code"], "会计科目编号直接引用 account_code。三级为6位；未来四级为8位，前6位必须是允许扩展的三级科目。没有明确业务映射时不得生成四级。"),
    record("TREE_REF_003003", "003003", ["account_name"], "会计科目名称必须与同一 account_code 联合投影，禁止单独生成或以名称反推编码。"),
    record("TREE_REF_003004", "003004", ["account_level"], "会计科目级次必须与同一 account_code 联合投影。当前种子为1至3级；未来获批四级科目固定返回4。"),
)


def build() -> dict[str, Any]:
    tree_meta = assert_ready()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if OUTPUT_DB.exists(): OUTPUT_DB.unlink()
    shutil.copy2(V1_DB, OUTPUT_DB)
    v1_sha = sha256_file(V1_DB)
    tree_sha = sha256_file(ACCOUNT_TREE_DB)
    conn = sqlite3.connect(OUTPUT_DB)
    try:
        meta = dict(conn.execute("SELECT meta_key,meta_value FROM extension_asset_meta").fetchall())
        if meta.get("publication_status") != "approved_partial_not_released":
            raise RuntimeError("源扩展V1不是已审批部分资产")
        if conn.execute("SELECT COUNT(*) FROM tree_reference_usage WHERE review_status<>'APPROVED'").fetchone()[0]:
            raise RuntimeError("源扩展V1存在未审批树引用")
        rows: list[dict[str, Any]] = []
        for item in RECORDS:
            quote, refs = refs_and_quote(conn, item["data_element_code"])
            rows.append({**item, "evidence_quote": quote, "source_refs_json": refs})
        columns = list(rows[0])
        conn.executemany(
            f"INSERT INTO tree_reference_usage ({','.join(columns)}) VALUES ({','.join('?' for _ in columns)})",
            [tuple(row[col] for col in columns) for row in rows],
        )
        now_utc = now()
        updates = {
            "asset_version": ASSET_VERSION,
            "publication_status": "approved_partial_not_released",
            "prior_extension_sqlite_sha256": v1_sha,
            "prior_extension_version": meta["asset_version"],
            "account_tree_asset_id": TREE_ASSET_ID,
            "account_tree_asset_status": tree_meta["asset_status"],
            "account_tree_sqlite_sha256": tree_sha,
            "account_tree_approval_ref": tree_meta["approval_ref"],
            "approval_ref": APPROVAL_REF,
            "approved_at_utc": now_utc,
            "not_in_scope": "单字段、表内、跨表、ODS等完整约束资产仍未建设；本版本只增加会计科目树及003002/003003/003004绑定。",
        }
        for key, value in updates.items():
            conn.execute("INSERT INTO extension_asset_meta(meta_key,meta_value) VALUES (?,?) ON CONFLICT(meta_key) DO UPDATE SET meta_value=excluded.meta_value", (key, value))
        conn.execute(
            "INSERT INTO extension_approval_log VALUES (?,?,?,?,?,?)",
            ("APPROVAL_CA_V0_2_0_ACCOUNT_TREE_BINDINGS", "TREE_REFERENCE_USAGE:003002,003003,003004", "APPROVED", APPROVAL_REF, now_utc,
             "用户批准无虚拟四级、三级同级扩展和按需8位四级扩展合同，并批准发布。"),
        )
        conn.commit()
    finally:
        conn.close()
    return validate(v1_sha, tree_sha)


def validate(v1_sha: str, tree_sha: str) -> dict[str, Any]:
    conn = sqlite3.connect(OUTPUT_DB)
    try:
        errors: list[str] = []
        if conn.execute("PRAGMA integrity_check").fetchone()[0] != "ok": errors.append("SQLite integrity_check失败")
        meta = dict(conn.execute("SELECT meta_key,meta_value FROM extension_asset_meta").fetchall())
        if meta.get("prior_extension_sqlite_sha256") != v1_sha: errors.append("扩展V1继承哈希不一致")
        if meta.get("account_tree_sqlite_sha256") != tree_sha: errors.append("科目树哈希不一致")
        count = conn.execute("SELECT COUNT(*) FROM tree_reference_usage").fetchone()[0]
        approved = conn.execute("SELECT COUNT(*) FROM tree_reference_usage WHERE review_status='APPROVED'").fetchone()[0]
        new = conn.execute("SELECT COUNT(*) FROM tree_reference_usage WHERE data_element_code IN ('003002','003003','003004') AND tree_asset_id=? AND dependency_status='READY'", (TREE_ASSET_ID,)).fetchone()[0]
        if count != 13 or approved != 13 or new != 3: errors.append("科目树引用记录不完整或未审批")
        if conn.execute("SELECT COUNT(*) FROM extension_approval_log WHERE approval_id='APPROVAL_CA_V0_2_0_ACCOUNT_TREE_BINDINGS'").fetchone()[0] != 1: errors.append("缺少V2审批日志")
        return {"status": "PASS" if not errors else "FAIL", "errors": errors, "tree_reference_total": count, "tree_reference_approved": approved, "new_account_tree_reference_count": new, "core_data_element_count": conn.execute("SELECT COUNT(*) FROM data_element").fetchone()[0]}
    finally:
        conn.close()


def write_excel(audit: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    conn = sqlite3.connect(OUTPUT_DB); conn.row_factory = sqlite3.Row
    try:
        refs = [dict(row) for row in conn.execute("SELECT * FROM tree_reference_usage ORDER BY data_element_code,tree_reference_id")]
        meta = [{"meta_key": k, "meta_value": v} for k,v in conn.execute("SELECT * FROM extension_asset_meta ORDER BY meta_key")]
        approval = [dict(row) for row in conn.execute("SELECT * FROM extension_approval_log ORDER BY approved_at_utc")]
    finally:
        conn.close()
    wb = Workbook(); wb.remove(wb.active)
    for title, rows in (("树引用", refs), ("资产元数据", meta), ("审批日志", approval), ("验收", [{"metric": k, "value": canonical(v) if isinstance(v,(dict,list)) else v} for k,v in audit.items()])):
        ws = wb.create_sheet(title); headers = list(rows[0]) if rows else ["empty"]; ws.append(headers)
        for row in rows: ws.append([row.get(header) for header in headers])
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.alignment = Alignment(horizontal="center")
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = min(max(max(len(str(c.value or "")) for c in col)+2, 12), 58)
        if rows:
            table = Table(displayName=f"T_{title}", ref=ws.dimensions); table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True); ws.add_table(table)
    wb.save(OUTPUT_DIR / "扩展数据元约束资产V2_人工审查.xlsx")


def write_design(audit: dict[str, Any]) -> None:
    (OUTPUT_DIR / "扩展数据元约束资产V2设计.md").write_text("\n".join([
        "# 扩展数据元约束资产 V2：机构部门树与会计科目树绑定", "",
        "状态：`approved_partial_not_released`", "",
        "V2 逐字节继承已审批的扩展 V1，并仅新增三个会计科目树引用：`003002 会计科目编号`、`003003 会计科目名称`、`003004 会计科目级次`。核心 EAST 资产未被改写。", "",
        "会计科目树不含虚拟四级节点。当前三级科目可记账且可扩展；四级仅在明确业务细分并经人工审批时新增，格式为8位、前6位为所属三级编码。编号、名称、级次始终由同一 `account_code` 投影。", "",
        f"人工审批依据：`{APPROVAL_REF}`", f"验收：`{canonical(audit)}`", "",
        "仍不发布为完整 V1 约束资产：单字段、表内、跨表、ODS 资产尚未完成。",
    ]) + "\n", encoding="utf-8")


def main() -> None:
    audit = build()
    if audit["status"] != "PASS": raise RuntimeError(canonical(audit))
    write_excel(audit); write_design(audit)
    (OUTPUT_DIR / "扩展资产V2构建审计.json").write_text(canonical(audit) + "\n", encoding="utf-8")
    print(canonical({"output": str(OUTPUT_DB), "sha256": sha256_file(OUTPUT_DB), "audit": audit}))


if __name__ == "__main__":
    main()
