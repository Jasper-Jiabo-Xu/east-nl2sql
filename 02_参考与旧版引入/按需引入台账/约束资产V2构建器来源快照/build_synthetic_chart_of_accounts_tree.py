#!/usr/bin/env python3
"""Build a review-only synthetic nationwide-bank chart of accounts.

This builder uses CoreBank *schema metadata* only: its public subject dictionary
shows that a usable general-ledger tree needs subject code, name, hierarchy
level, debit/credit direction, type, summary/control dimensions and status.
It never opens ``corebank.db`` and contains no copied legacy subject record.

The account catalogue below is deliberately synthetic.  It is a broad, stable
banking accounting skeleton for EAST data generation, not an official Chinese
accounting-standard chart and not a real bank's account setup.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[2]
PHASE = ROOT / "kb/working/20260727_000_goal_mode_v1/构建过程层/04_字段多字段与对象明细状态提取-V2"
OUTPUT_DIR = PHASE / "extension_assets_v1/20260805_会计科目树V1_人工审查"

ASSET_VERSION = "extension-chart-of-accounts-tree-v1-20260805"
ASSET_ID = "SYNTHETIC_CHART_OF_ACCOUNTS_TREE_V1"
COREBANK_SCHEMA_REF = "SRC_COREBANK_SCHEMA:info_analyzed_non_empty_datetime_adj.xlsx!bak_公共科目字典"
EAST_REFS = {
    "003002": "SRC_ATT3:数据元说明!F78",
    "003003": "SRC_ATT3:数据元说明!F79",
    "003004": "SRC_ATT3:数据元说明!F80",
}
DESIGN_REF = "SRC_EXTENSION_DESIGN:会计科目树V1"


def canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


# Code contract: XX / XXXX / XXXXXX / XXXXXXXX.  A first-level class has two
# digits and a second-level group reserves an entire two-digit suffix.  Seeded
# level-3 accounts use only 10, 20, ... 90, intentionally leaving sibling
# insertion positions.  A future level-4 detail subject appends two digits to
# an approved, expandable level-3 code.  The current seeds intentionally have
# no artificial fourth-level children.
CLASSES: tuple[tuple[str, str, str, str], ...] = (
    ("10", "资产类", "ASSET", "DEBIT"),
    ("20", "负债类", "LIABILITY", "CREDIT"),
    ("30", "所有者权益类", "EQUITY", "CREDIT"),
    ("40", "共同类", "COMMON", "BOTH"),
    ("50", "成本类", "COST", "DEBIT"),
    ("60", "损益收入类", "INCOME", "CREDIT"),
    ("70", "损益费用类", "EXPENSE", "DEBIT"),
    ("80", "表外及备查类", "OFF_BALANCE", "BOTH"),
)

# (class code, group suffix, group name, account nature, normal balance,
#  [(leaf suffix, leaf name, reporting domain, account role), ...])
GROUPS: tuple[tuple[str, str, str, str, str, tuple[tuple[str, str, str, str], ...]], ...] = (
    ("10", "01", "现金及中央银行款项", "ASSET", "DEBIT", (
        ("10", "库存现金", "CASH_MANAGEMENT", "OPERATING"),
        ("20", "在途现金", "CASH_MANAGEMENT", "OPERATING"),
        ("30", "自助设备现金", "CASH_MANAGEMENT", "OPERATING"),
        ("40", "存放中央银行款项", "CENTRAL_BANK", "OPERATING"),
        ("50", "法定存款准备金", "CENTRAL_BANK", "OPERATING"),
        ("60", "超额准备金", "CENTRAL_BANK", "OPERATING"),
        ("70", "中央银行专项资金", "CENTRAL_BANK", "OPERATING"),
    )),
    ("10", "02", "同业及清算资金", "ASSET", "DEBIT", (
        ("10", "存放同业款项", "INTERBANK", "OPERATING"),
        ("20", "拆放同业款项", "INTERBANK", "OPERATING"),
        ("30", "买入返售金融资产", "INTERBANK", "OPERATING"),
        ("40", "同业清算应收款", "PAYMENT_SETTLEMENT", "OPERATING"),
        ("50", "支付清算在途资金", "PAYMENT_SETTLEMENT", "OPERATING"),
        ("60", "贵金属存放款项", "TREASURY", "OPERATING"),
    )),
    ("10", "03", "贷款及垫款", "ASSET", "DEBIT", (
        ("10", "公司流动资金贷款", "CORPORATE_CREDIT", "OPERATING"),
        ("20", "公司固定资产贷款", "CORPORATE_CREDIT", "OPERATING"),
        ("30", "个人住房贷款", "RETAIL_CREDIT", "OPERATING"),
        ("40", "个人消费贷款", "RETAIL_CREDIT", "OPERATING"),
        ("50", "个人经营贷款", "RETAIL_CREDIT", "OPERATING"),
        ("60", "信用卡透支", "CARD", "OPERATING"),
        ("70", "贴现及转贴现资产", "BILL_FINANCE", "OPERATING"),
        ("80", "融资租赁应收款", "LEASE", "OPERATING"),
        ("90", "贷款减值准备", "CREDIT_IMPAIRMENT", "CONTRA"),
    )),
    ("10", "04", "金融投资及衍生金融资产", "ASSET", "DEBIT", (
        ("10", "交易性债务工具投资", "MARKET_RISK", "OPERATING"),
        ("20", "交易性权益工具投资", "MARKET_RISK", "OPERATING"),
        ("30", "以公允价值计量且其变动计入其他综合收益的债务工具", "INVESTMENT", "OPERATING"),
        ("40", "以公允价值计量且其变动计入其他综合收益的权益工具", "INVESTMENT", "OPERATING"),
        ("50", "以摊余成本计量的债务工具投资", "INVESTMENT", "OPERATING"),
        ("60", "衍生金融资产", "DERIVATIVE", "OPERATING"),
        ("70", "金融投资减值准备", "CREDIT_IMPAIRMENT", "CONTRA"),
    )),
    ("10", "05", "长期股权及其他权益性投资", "ASSET", "DEBIT", (
        ("10", "长期股权投资", "EQUITY_INVESTMENT", "OPERATING"),
        ("20", "联营企业及合营企业投资", "EQUITY_INVESTMENT", "OPERATING"),
        ("30", "战略性权益投资", "EQUITY_INVESTMENT", "OPERATING"),
        ("40", "其他权益工具投资", "EQUITY_INVESTMENT", "OPERATING"),
    )),
    ("10", "06", "固定资产及使用权资产", "ASSET", "DEBIT", (
        ("10", "固定资产原值", "FIXED_ASSET", "OPERATING"),
        ("20", "累计折旧", "FIXED_ASSET", "CONTRA"),
        ("30", "在建工程", "FIXED_ASSET", "OPERATING"),
        ("40", "投资性房地产", "FIXED_ASSET", "OPERATING"),
        ("50", "使用权资产", "LEASE", "OPERATING"),
        ("60", "使用权资产累计折旧", "LEASE", "CONTRA"),
    )),
    ("10", "07", "无形资产及递延资产", "ASSET", "DEBIT", (
        ("10", "无形资产", "INTANGIBLE_ASSET", "OPERATING"),
        ("20", "累计摊销", "INTANGIBLE_ASSET", "CONTRA"),
        ("30", "递延所得税资产", "TAX", "OPERATING"),
        ("40", "待摊费用", "PREPAID", "OPERATING"),
        ("50", "长期待摊费用", "PREPAID", "OPERATING"),
    )),
    ("10", "08", "其他资产", "ASSET", "DEBIT", (
        ("10", "应收利息", "ACCRUAL", "OPERATING"),
        ("20", "其他应收款", "OTHER_RECEIVABLE", "OPERATING"),
        ("30", "抵债资产", "COLLATERAL", "OPERATING"),
        ("40", "待处理财产损溢", "SUSPENSE", "OPERATING"),
        ("50", "其他资产", "OTHER_ASSET", "OPERATING"),
    )),
    ("20", "01", "客户存款", "LIABILITY", "CREDIT", (
        ("10", "个人活期存款", "RETAIL_DEPOSIT", "OPERATING"),
        ("20", "个人定期存款", "RETAIL_DEPOSIT", "OPERATING"),
        ("30", "单位活期存款", "CORPORATE_DEPOSIT", "OPERATING"),
        ("40", "单位定期存款", "CORPORATE_DEPOSIT", "OPERATING"),
        ("50", "保证金存款", "DEPOSIT", "OPERATING"),
        ("60", "协议存款", "DEPOSIT", "OPERATING"),
        ("70", "结构性存款", "DEPOSIT", "OPERATING"),
        ("80", "应付存款利息", "ACCRUAL", "OPERATING"),
    )),
    ("20", "02", "同业负债及向中央银行借款", "LIABILITY", "CREDIT", (
        ("10", "同业存放款项", "INTERBANK", "OPERATING"),
        ("20", "同业拆入资金", "INTERBANK", "OPERATING"),
        ("30", "卖出回购金融资产款", "INTERBANK", "OPERATING"),
        ("40", "向中央银行借款", "CENTRAL_BANK", "OPERATING"),
        ("50", "再贷款资金", "CENTRAL_BANK", "OPERATING"),
        ("60", "非银行金融机构拆入资金", "INTERBANK", "OPERATING"),
    )),
    ("20", "03", "应付债券及其他融资", "LIABILITY", "CREDIT", (
        ("10", "金融债券", "BOND_FINANCING", "OPERATING"),
        ("20", "同业存单", "BOND_FINANCING", "OPERATING"),
        ("30", "二级资本债", "BOND_FINANCING", "OPERATING"),
        ("40", "可转换债券", "BOND_FINANCING", "OPERATING"),
        ("50", "其他应付债务融资工具", "BOND_FINANCING", "OPERATING"),
        ("60", "应付债券利息", "ACCRUAL", "OPERATING"),
    )),
    ("20", "04", "交易性及衍生金融负债", "LIABILITY", "CREDIT", (
        ("10", "交易性金融负债", "MARKET_RISK", "OPERATING"),
        ("20", "衍生金融负债", "DERIVATIVE", "OPERATING"),
        ("30", "卖空金融资产款", "MARKET_RISK", "OPERATING"),
        ("40", "指定为以公允价值计量且其变动计入当期损益的金融负债", "MARKET_RISK", "OPERATING"),
    )),
    ("20", "05", "应付款项及职工薪酬", "LIABILITY", "CREDIT", (
        ("10", "应付职工薪酬", "HR", "OPERATING"),
        ("20", "应交税费", "TAX", "OPERATING"),
        ("30", "应付手续费及佣金", "ACCRUAL", "OPERATING"),
        ("40", "应付清算款", "PAYMENT_SETTLEMENT", "OPERATING"),
        ("50", "其他应付款", "OTHER_PAYABLE", "OPERATING"),
        ("60", "预计负债", "PROVISION", "OPERATING"),
    )),
    ("20", "06", "租赁及递延负债", "LIABILITY", "CREDIT", (
        ("10", "租赁负债", "LEASE", "OPERATING"),
        ("20", "递延所得税负债", "TAX", "OPERATING"),
        ("30", "递延收益", "DEFERRED_INCOME", "OPERATING"),
        ("40", "其他负债", "OTHER_LIABILITY", "OPERATING"),
    )),
    ("30", "01", "实收资本及资本公积", "EQUITY", "CREDIT", (
        ("10", "实收资本", "CAPITAL", "OPERATING"),
        ("20", "资本公积", "CAPITAL", "OPERATING"),
        ("30", "其他权益工具", "CAPITAL", "OPERATING"),
    )),
    ("30", "02", "储备及未分配利润", "EQUITY", "CREDIT", (
        ("10", "盈余公积", "RESERVE", "OPERATING"),
        ("20", "一般风险准备", "RESERVE", "OPERATING"),
        ("30", "其他综合收益", "OCI", "OPERATING"),
        ("40", "本年利润", "PROFIT", "OPERATING"),
        ("50", "利润分配", "PROFIT", "OPERATING"),
        ("60", "未分配利润", "PROFIT", "OPERATING"),
    )),
    ("30", "03", "少数股东权益", "EQUITY", "CREDIT", (
        ("10", "少数股东权益", "CONSOLIDATION", "OPERATING"),
        ("20", "归属于母公司股东的其他权益", "CONSOLIDATION", "OPERATING"),
    )),
    ("40", "01", "资金清算共同类", "COMMON", "BOTH", (
        ("10", "系统内往来", "INTERNAL_SETTLEMENT", "OPERATING"),
        ("20", "同城票据清算", "PAYMENT_SETTLEMENT", "OPERATING"),
        ("30", "跨行支付清算", "PAYMENT_SETTLEMENT", "OPERATING"),
        ("40", "外汇买卖清算", "FX_SETTLEMENT", "OPERATING"),
        ("50", "待清算资金", "PAYMENT_SETTLEMENT", "OPERATING"),
    )),
    ("40", "02", "内部转移及过渡共同类", "COMMON", "BOTH", (
        ("10", "内部资金转移", "FTP", "OPERATING"),
        ("20", "内部损益结转", "INTERNAL_SETTLEMENT", "OPERATING"),
        ("30", "跨机构往来", "INTERNAL_SETTLEMENT", "OPERATING"),
        ("40", "暂收暂付款", "SUSPENSE", "OPERATING"),
    )),
    ("50", "01", "资金成本", "COST", "DEBIT", (
        ("10", "内部资金转移成本", "FTP", "OPERATING"),
        ("20", "资金筹集成本", "FUNDING", "OPERATING"),
        ("30", "发行融资工具成本", "FUNDING", "OPERATING"),
    )),
    ("50", "02", "业务及项目成本", "COST", "DEBIT", (
        ("10", "受托及代理业务成本", "AGENCY", "OPERATING"),
        ("20", "银行卡业务成本", "CARD", "OPERATING"),
        ("30", "金融市场业务成本", "TREASURY", "OPERATING"),
    )),
    ("60", "01", "利息收入", "INCOME", "CREDIT", (
        ("10", "发放贷款及垫款利息收入", "INTEREST_INCOME", "OPERATING"),
        ("20", "金融投资利息收入", "INTEREST_INCOME", "OPERATING"),
        ("30", "存放及拆放同业利息收入", "INTEREST_INCOME", "OPERATING"),
        ("40", "融资租赁利息收入", "INTEREST_INCOME", "OPERATING"),
        ("50", "信用卡利息收入", "INTEREST_INCOME", "OPERATING"),
    )),
    ("60", "02", "手续费及佣金收入", "INCOME", "CREDIT", (
        ("10", "结算与清算手续费收入", "FEE_INCOME", "OPERATING"),
        ("20", "银行卡手续费收入", "FEE_INCOME", "OPERATING"),
        ("30", "代理及托管手续费收入", "FEE_INCOME", "OPERATING"),
        ("40", "理财及投资顾问手续费收入", "FEE_INCOME", "OPERATING"),
        ("50", "担保及承诺手续费收入", "FEE_INCOME", "OPERATING"),
    )),
    ("60", "03", "投资及公允价值收益", "INCOME", "CREDIT", (
        ("10", "投资收益", "INVESTMENT_INCOME", "OPERATING"),
        ("20", "公允价值变动收益", "FAIR_VALUE", "OPERATING"),
        ("30", "汇兑收益", "FX", "OPERATING"),
        ("40", "资产处置收益", "ASSET_DISPOSAL", "OPERATING"),
        ("50", "其他综合收益结转", "OCI", "OPERATING"),
    )),
    ("60", "04", "其他业务收入", "INCOME", "CREDIT", (
        ("10", "其他业务收入", "OTHER_INCOME", "OPERATING"),
        ("20", "营业外收入", "NON_OPERATING_INCOME", "OPERATING"),
        ("30", "政府补助收益", "GOVERNMENT_GRANT", "OPERATING"),
        ("40", "营业外损益调整收入", "NON_OPERATING_INCOME", "OPERATING"),
    )),
    ("70", "01", "利息支出", "EXPENSE", "DEBIT", (
        ("10", "吸收存款利息支出", "INTEREST_EXPENSE", "OPERATING"),
        ("20", "同业及中央银行融资利息支出", "INTEREST_EXPENSE", "OPERATING"),
        ("30", "应付债券利息支出", "INTEREST_EXPENSE", "OPERATING"),
        ("40", "租赁负债利息支出", "INTEREST_EXPENSE", "OPERATING"),
    )),
    ("70", "02", "手续费及佣金支出", "EXPENSE", "DEBIT", (
        ("10", "结算与清算手续费支出", "FEE_EXPENSE", "OPERATING"),
        ("20", "银行卡手续费支出", "FEE_EXPENSE", "OPERATING"),
        ("30", "代理及托管手续费支出", "FEE_EXPENSE", "OPERATING"),
        ("40", "业务服务费支出", "FEE_EXPENSE", "OPERATING"),
    )),
    ("70", "03", "信用及资产减值损失", "EXPENSE", "DEBIT", (
        ("10", "贷款信用减值损失", "CREDIT_IMPAIRMENT", "OPERATING"),
        ("20", "金融投资信用减值损失", "CREDIT_IMPAIRMENT", "OPERATING"),
        ("30", "其他应收款信用减值损失", "CREDIT_IMPAIRMENT", "OPERATING"),
        ("40", "固定资产减值损失", "ASSET_IMPAIRMENT", "OPERATING"),
        ("50", "抵债资产减值损失", "ASSET_IMPAIRMENT", "OPERATING"),
    )),
    ("70", "04", "运营及管理费用", "EXPENSE", "DEBIT", (
        ("10", "职工薪酬", "PERSONNEL", "OPERATING"),
        ("20", "业务及行政管理费", "ADMINISTRATIVE", "OPERATING"),
        ("30", "折旧费", "DEPRECIATION", "OPERATING"),
        ("40", "无形资产摊销", "AMORTIZATION", "OPERATING"),
        ("50", "租赁及物业费用", "PROPERTY", "OPERATING"),
        ("60", "信息科技费用", "TECHNOLOGY", "OPERATING"),
    )),
    ("70", "05", "税费及其他损失", "EXPENSE", "DEBIT", (
        ("10", "税金及附加", "TAX", "OPERATING"),
        ("20", "营业外支出", "NON_OPERATING_EXPENSE", "OPERATING"),
        ("30", "资产处置损失", "ASSET_DISPOSAL", "OPERATING"),
        ("40", "公允价值变动损失", "FAIR_VALUE", "OPERATING"),
        ("50", "汇兑损失", "FX", "OPERATING"),
        ("60", "其他业务支出", "OTHER_EXPENSE", "OPERATING"),
    )),
    ("80", "01", "授信承诺及未使用额度", "OFF_BALANCE", "BOTH", (
        ("10", "未使用贷款承诺", "CREDIT_COMMITMENT", "MEMORANDUM"),
        ("20", "信用卡未使用额度", "CREDIT_COMMITMENT", "MEMORANDUM"),
        ("30", "融资额度承诺", "CREDIT_COMMITMENT", "MEMORANDUM"),
        ("40", "已用授信额度备查", "CREDIT_COMMITMENT", "MEMORANDUM"),
    )),
    ("80", "02", "担保及承兑", "OFF_BALANCE", "BOTH", (
        ("10", "银行承兑汇票", "GUARANTEE", "MEMORANDUM"),
        ("20", "保函及备用信用证", "GUARANTEE", "MEMORANDUM"),
        ("30", "信用证及进口押汇承诺", "GUARANTEE", "MEMORANDUM"),
        ("40", "担保合同余额", "GUARANTEE", "MEMORANDUM"),
    )),
    ("80", "03", "衍生工具名义本金", "OFF_BALANCE", "BOTH", (
        ("10", "利率衍生工具名义本金", "DERIVATIVE", "MEMORANDUM"),
        ("20", "汇率衍生工具名义本金", "DERIVATIVE", "MEMORANDUM"),
        ("30", "信用衍生工具名义本金", "DERIVATIVE", "MEMORANDUM"),
        ("40", "商品及其他衍生工具名义本金", "DERIVATIVE", "MEMORANDUM"),
    )),
    ("80", "04", "受托及代理资产负债", "OFF_BALANCE", "BOTH", (
        ("10", "委托贷款及委托投资", "TRUSTEE", "MEMORANDUM"),
        ("20", "受托理财资产", "TRUSTEE", "MEMORANDUM"),
        ("30", "托管资产", "CUSTODY", "MEMORANDUM"),
        ("40", "代理收付及代销资产", "AGENCY", "MEMORANDUM"),
    )),
    ("80", "05", "备查及风险暴露", "OFF_BALANCE", "BOTH", (
        ("10", "抵质押品登记价值", "COLLATERAL", "MEMORANDUM"),
        ("20", "已核销资产备查", "WRITE_OFF", "MEMORANDUM"),
        ("30", "诉讼及或有事项备查", "CONTINGENCY", "MEMORANDUM"),
        ("40", "其他风险暴露备查", "RISK_EXPOSURE", "MEMORANDUM"),
    )),
)


def node(
    code: str, parent: str | None, level: int, name: str, account_class: str,
    nature: str, normal_balance: str, postable: bool, role: str, domain: str,
) -> dict[str, Any]:
    return {
        "account_code": code,
        "parent_account_code": parent,
        "account_level": level,
        "account_name": name,
        "full_name": None,
        "account_class_code": account_class,
        "account_nature": nature,
        "normal_balance": normal_balance,
        "is_summary": 0 if postable else 1,
        "is_postable": 1 if postable else 0,
        "posting_policy": "POSTABLE_AND_EXPANDABLE" if postable and level == 3 else ("POSTABLE" if postable else "SUMMARY_ONLY"),
        "account_role": role,
        "business_domain": domain,
        "status_code": "ACTIVE",
        "effective_from": "2020-01-01",
        "effective_to": None,
        "is_synthetic": 1,
        "source_class": "SYNTHETIC_EXTENSION_DESIGN",
        "review_status": "PENDING_HUMAN_REVIEW",
        "source_refs_json": canonical([COREBANK_SCHEMA_REF, *EAST_REFS.values(), DESIGN_REF]),
        "attributes_json": canonical({
            "code_contract": "XX/XXXX/XXXXXX; leaf suffixes reserve 9 insertion positions per group",
            "legacy_data_imported": False,
            "notes": "合成科目树；不得视为真实银行科目、法定科目表或直接写账凭证依据。",
        }),
    }


def assign_full_names(nodes: list[dict[str, Any]]) -> None:
    by_code = {item["account_code"]: item for item in nodes}
    memo: dict[str, str] = {}

    def resolve(code: str) -> str:
        if code in memo:
            return memo[code]
        current = by_code[code]
        parent = current["parent_account_code"]
        value = current["account_name"] if parent is None else f"{resolve(parent)} / {current['account_name']}"
        memo[code] = value
        return value

    for item in nodes:
        item["full_name"] = resolve(str(item["account_code"]))


def build_nodes() -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    class_by_code = {code: (name, nature, balance) for code, name, nature, balance in CLASSES}
    nodes: list[dict[str, Any]] = []
    relations: list[dict[str, str]] = []
    for class_code, class_name, nature, balance in CLASSES:
        nodes.append(node(class_code, None, 1, class_name, class_code, nature, balance, False, "SUMMARY", "ALL"))
    for class_code, group_suffix, group_name, nature, balance, leaves in GROUPS:
        if class_code not in class_by_code:
            raise RuntimeError(f"未知科目大类：{class_code}")
        group_code = f"{class_code}{group_suffix}"
        nodes.append(node(group_code, class_code, 2, group_name, class_code, nature, balance, False, "SUMMARY", "ALL"))
        relations.append({"from_account_code": group_code, "to_account_code": class_code, "relation_type": "TREE_PARENT"})
        for leaf_suffix, leaf_name, domain, role in leaves:
            leaf_code = f"{group_code}{leaf_suffix}"
            leaf_balance = balance
            leaf_nature = nature
            if role == "CONTRA":
                leaf_balance = "CREDIT" if nature == "ASSET" else "DEBIT"
                leaf_nature = f"{nature}_CONTRA"
            nodes.append(node(leaf_code, group_code, 3, leaf_name, class_code, leaf_nature, leaf_balance, True, role, domain))
            relations.append({"from_account_code": leaf_code, "to_account_code": group_code, "relation_type": "TREE_PARENT"})
    assign_full_names(nodes)
    return nodes, relations


def create_db(path: Path, nodes: list[dict[str, Any]], relations: list[dict[str, str]]) -> None:
    if path.exists():
        path.unlink()
    conn = sqlite3.connect(path)
    try:
        conn.execute("PRAGMA foreign_keys=ON")
        conn.executescript(
            """
            CREATE TABLE asset_meta (asset_key TEXT PRIMARY KEY, asset_value TEXT NOT NULL);
            CREATE TABLE account_node (
                account_code TEXT PRIMARY KEY CHECK(length(account_code) IN (2,4,6,8)),
                parent_account_code TEXT REFERENCES account_node(account_code),
                account_level INTEGER NOT NULL CHECK(account_level BETWEEN 1 AND 4),
                account_name TEXT NOT NULL,
                full_name TEXT NOT NULL UNIQUE,
                account_class_code TEXT NOT NULL,
                account_nature TEXT NOT NULL,
                normal_balance TEXT NOT NULL CHECK(normal_balance IN ('DEBIT','CREDIT','BOTH')),
                is_summary INTEGER NOT NULL CHECK(is_summary IN (0,1)),
                is_postable INTEGER NOT NULL CHECK(is_postable IN (0,1)),
                posting_policy TEXT NOT NULL CHECK(posting_policy IN ('SUMMARY_ONLY','POSTABLE','POSTABLE_AND_EXPANDABLE')),
                account_role TEXT NOT NULL CHECK(account_role IN ('SUMMARY','OPERATING','CONTRA','MEMORANDUM')),
                business_domain TEXT NOT NULL,
                status_code TEXT NOT NULL CHECK(status_code='ACTIVE'),
                effective_from TEXT NOT NULL,
                effective_to TEXT,
                is_synthetic INTEGER NOT NULL CHECK(is_synthetic=1),
                source_class TEXT NOT NULL CHECK(source_class='SYNTHETIC_EXTENSION_DESIGN'),
                review_status TEXT NOT NULL CHECK(review_status IN ('PENDING_HUMAN_REVIEW','APPROVED')),
                source_refs_json TEXT NOT NULL,
                attributes_json TEXT NOT NULL,
                CHECK((is_summary=1 AND is_postable=0 AND posting_policy='SUMMARY_ONLY') OR
                      (is_summary=0 AND is_postable=1 AND posting_policy IN ('POSTABLE','POSTABLE_AND_EXPANDABLE'))),
                CHECK((account_level=1 AND parent_account_code IS NULL) OR (account_level>1 AND parent_account_code IS NOT NULL))
            );
            CREATE TABLE account_relation (
                from_account_code TEXT NOT NULL REFERENCES account_node(account_code),
                to_account_code TEXT NOT NULL REFERENCES account_node(account_code),
                relation_type TEXT NOT NULL CHECK(relation_type='TREE_PARENT'),
                PRIMARY KEY(from_account_code,to_account_code,relation_type)
            );
            CREATE TABLE east_data_element_binding_plan (
                data_element_code TEXT PRIMARY KEY CHECK(data_element_code IN ('003002','003003','003004')),
                data_element_name TEXT NOT NULL,
                target_table TEXT NOT NULL CHECK(target_table='account_node'),
                lookup_key_column TEXT NOT NULL,
                value_projection_json TEXT NOT NULL,
                binding_mode TEXT NOT NULL CHECK(binding_mode='DIRECT_VALUE'),
                applicability_condition_json TEXT NOT NULL,
                fallback_policy TEXT NOT NULL CHECK(fallback_policy='NO_FALLBACK_MANUAL_REVIEW'),
                dependency_status TEXT NOT NULL CHECK(dependency_status IN ('PENDING_EXTENSION_ASSET_BINDING','READY')),
                source_refs_json TEXT NOT NULL,
                review_status TEXT NOT NULL CHECK(review_status IN ('PENDING_HUMAN_REVIEW','APPROVED')),
                notes TEXT NOT NULL
            );
            CREATE INDEX idx_account_node_parent ON account_node(parent_account_code);
            CREATE INDEX idx_account_node_postable ON account_node(is_postable, status_code);
            CREATE INDEX idx_account_node_domain ON account_node(business_domain);
            """
        )
        cols = list(nodes[0])
        conn.executemany(
            f"INSERT INTO account_node ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})",
            [tuple(item[column] for column in cols) for item in nodes],
        )
        conn.executemany(
            "INSERT INTO account_relation VALUES (:from_account_code,:to_account_code,:relation_type)", relations,
        )
        binding_rows = (
            ("003002", "会计科目编号", "account_code", canonical(["account_code"]), "总账会计科目引用 is_postable=1 的节点。当前三级为6位；四级不设虚拟值，未来仅能以8位代码新增，且前6位为已批准的 POSTABLE_AND_EXPANDABLE 三级代码。"),
            ("003003", "会计科目名称", "account_code", canonical(["account_name"]), "由同一 account_code 投影 account_name；禁止独立生成导致名称与编号不一致。"),
            ("003004", "会计科目级次", "account_code", canonical(["account_level"]), "由同一 account_code 投影 account_level；当前合同层级为 1/2/3。"),
        )
        for code, name, lookup, projection, notes in binding_rows:
            conn.execute(
                """INSERT INTO east_data_element_binding_plan VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (code, name, "account_node", lookup, projection, "DIRECT_VALUE", canonical({"account_status": "ACTIVE"}),
                 "NO_FALLBACK_MANUAL_REVIEW", "PENDING_EXTENSION_ASSET_BINDING", canonical([EAST_REFS[code], DESIGN_REF]),
                 "PENDING_HUMAN_REVIEW", notes),
            )
        meta = {
            "asset_id": ASSET_ID,
            "asset_version": ASSET_VERSION,
            "asset_status": "review_only_not_released",
            "tree_type": "SYNTHETIC_BANK_CHART_OF_ACCOUNTS",
            "code_contract": "XX/XXXX/XXXXXX/XXXXXXXX; each level-2 group reserves level-3 suffix positions 01-99, only 10-90 are seeded; level-4 appends two digits to an approved expandable level-3 account",
            "fourth_level_policy": "no_virtual_level4_nodes; a level-4 code is created only for an approved business distinction; it must have a six-digit POSTABLE_AND_EXPANDABLE parent and an eight-digit code with that parent as prefix",
            "effective_scope": "EAST虚拟全国银行；仅供约束与生成设计，不代表任何真实银行或法定科目表",
            "corebank_usage_boundary": "schema_metadata_only; no corebank database or account record opened/copied",
            "source_refs_json": canonical([COREBANK_SCHEMA_REF, *EAST_REFS.values(), DESIGN_REF]),
            "built_at_utc": now(),
            "not_publishable_until": "全部节点及003002/003003/003004绑定计划经人工审批；不得混入真实CoreBank或真实银行科目配置。",
        }
        conn.executemany("INSERT INTO asset_meta VALUES (?,?)", list(meta.items()))
        conn.commit()
    finally:
        conn.close()


def validate(path: Path) -> dict[str, Any]:
    conn = sqlite3.connect(path)
    try:
        errors: list[str] = []
        if conn.execute("PRAGMA integrity_check").fetchone()[0] != "ok":
            errors.append("SQLite integrity_check 失败")
        nodes = conn.execute("SELECT COUNT(*) FROM account_node").fetchone()[0]
        relations = conn.execute("SELECT COUNT(*) FROM account_relation").fetchone()[0]
        leaves = conn.execute("SELECT COUNT(*) FROM account_node WHERE is_postable=1").fetchone()[0]
        summaries = conn.execute("SELECT COUNT(*) FROM account_node WHERE is_summary=1").fetchone()[0]
        roots = conn.execute("SELECT COUNT(*) FROM account_node WHERE account_level=1 AND parent_account_code IS NULL").fetchone()[0]
        invalid_parent = conn.execute("""
            SELECT COUNT(*) FROM account_node child JOIN account_node parent ON child.parent_account_code=parent.account_code
             WHERE child.account_level<>parent.account_level+1
        """).fetchone()[0]
        non_seed_posting = conn.execute("SELECT COUNT(*) FROM account_node WHERE account_level<3 AND is_postable=1").fetchone()[0]
        seeded_level3_not_expandable = conn.execute("SELECT COUNT(*) FROM account_node WHERE account_level=3 AND posting_policy<>'POSTABLE_AND_EXPANDABLE'").fetchone()[0]
        bad_reserved = conn.execute("SELECT COUNT(*) FROM account_node WHERE account_level=3 AND substr(account_code,5,2) NOT IN ('10','20','30','40','50','60','70','80','90')").fetchone()[0]
        invalid_level4 = conn.execute("""
            SELECT COUNT(*) FROM account_node child LEFT JOIN account_node parent ON child.parent_account_code=parent.account_code
             WHERE child.account_level=4 AND (length(child.account_code)<>8 OR length(parent.account_code)<>6
                OR substr(child.account_code,1,6)<>parent.account_code OR parent.posting_policy<>'POSTABLE_AND_EXPANDABLE')
        """).fetchone()[0]
        duplicate_name = conn.execute("SELECT COUNT(*) FROM (SELECT full_name,COUNT(*) n FROM account_node GROUP BY full_name HAVING n>1)").fetchone()[0]
        if roots != len(CLASSES): errors.append("一级科目数量不完整")
        if nodes != relations + roots: errors.append("树边数不满足 node_count = relation_count + root_count")
        if invalid_parent: errors.append(f"存在{invalid_parent}条跨级父子关系")
        if non_seed_posting or seeded_level3_not_expandable: errors.append("当前三级种子科目的记账/可扩展策略不一致")
        if bad_reserved: errors.append("末级编码未遵循预留位合同")
        if invalid_level4: errors.append("四级科目不满足前六位父级及可扩展父级合同")
        if duplicate_name: errors.append("完整名称不唯一")
        bindings = conn.execute("SELECT COUNT(*) FROM east_data_element_binding_plan").fetchone()[0]
        if bindings != 3: errors.append("EAST数据元绑定计划不完整")
        return {
            "status": "PASS" if not errors else "FAIL", "errors": errors,
            "node_count": nodes, "relation_count": relations, "root_count": roots,
            "summary_count": summaries, "postable_leaf_count": leaves,
            "class_distribution": dict(conn.execute("SELECT account_class_code,COUNT(*) FROM account_node GROUP BY account_class_code ORDER BY account_class_code").fetchall()),
            "nature_distribution": dict(conn.execute("SELECT account_nature,COUNT(*) FROM account_node GROUP BY account_nature ORDER BY account_nature").fetchall()),
            "binding_plan_count": bindings,
        }
    finally:
        conn.close()


def write_design(path: Path, audit: dict[str, Any], db_sha: str, asset_status: str = "review_only_not_released", approval_ref: str | None = None) -> None:
    lines = [
        "# 会计科目树 V1（人工审查版）", "",
        f"状态：`{asset_status}`  ", f"版本：`{ASSET_VERSION}`", "",
        "## 边界", "",
        "- 本树是 EAST 虚拟全国银行的合成总账科目骨架，不是法定科目表或任何真实银行的科目配置。",
        "- CoreBank 只提供公共科目字典的**字段结构**参照：科目号、名称、级别、借贷标志、类型与汇总/控制维度；未打开 CoreBank 数据库、未读取或复制任何科目记录。",
        "- 当前仅覆盖 `003002 会计科目编号`、`003003 会计科目名称`、`003004 会计科目级次` 的候选引用计划；尚未写入扩展数据元约束资产。", "",
        "## 编码合同", "",
        "`account_code` 固定为 `XX / XXXX / XXXXXX / XXXXXXXX`：两位一级大类、四位二级业务组、六位三级科目、八位四级明细科目。当前只投放三级真实科目，三级末两位仅使用 `10、20、…、90`，每个二级组保留其余位置以便补充同级科目。", "",
        "不创建虚拟四级节点。未来只有发现明确业务细分时，才可增加八位科目；其前六位必须等于一个 `POSTABLE_AND_EXPANDABLE` 的三级科目。三级科目在增加四级后仍可承接未细分业务，已有具体映射的业务优先使用四级。", "",
        "运行时：业务明细只允许引用 `is_postable=1` 的节点；编号、名称、级次必须由同一 `account_code` 联合投影，禁止分别生成。", "",
        "## 人工审查重点", "",
        "- 资产、负债、权益、共同、成本、收入、费用、表外八大类及业务覆盖是否适合 EAST 后续题目。",
        "- 末级科目的命名、借贷方向、冲减科目与表外备查科目是否需要调整。",
        "- 是否接受四层 `XX/XXXX/XXXXXX/XXXXXXXX` 编码，以及三级可记账且可扩展、不设虚拟四级节点的策略。", "",
        "## 构建验收", "",
        f"- SHA-256：`{db_sha}`", f"- 节点：{audit['node_count']}；父子关系：{audit['relation_count']}；可记账末级：{audit['postable_leaf_count']}；汇总节点：{audit['summary_count']}",
        f"- 大类分布：`{canonical(audit['class_distribution'])}`", f"- 结构校验：`{audit['status']}`", "",
        "## 不发布条件", "",
        "- 任一节点或三个 EAST 数据元绑定计划仍为 `PENDING_HUMAN_REVIEW`；",
        "- 科目编码、层级、方向、可记账范围未获人工批准；",
        "- 发现真实 CoreBank 或真实银行的科目记录、余额、账户或配置信息被写入。",
    ]
    if approval_ref:
        lines.insert(5, f"人工审批依据：`{approval_ref}`")
        lines.insert(6, "")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_excel(path: Path, db_path: Path, audit: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        node_rows = [dict(row) for row in conn.execute("SELECT * FROM account_node ORDER BY account_code")]
        binding_rows = [dict(row) for row in conn.execute("SELECT * FROM east_data_element_binding_plan ORDER BY data_element_code")]
        meta_rows = [{"asset_key": k, "asset_value": v} for k, v in conn.execute("SELECT * FROM asset_meta ORDER BY asset_key")]
    finally:
        conn.close()
    wb = Workbook(); wb.remove(wb.active)
    sheets = (("科目树", node_rows), ("EAST绑定计划", binding_rows), ("资产元数据", meta_rows), ("验收", [{"metric": k, "value": canonical(v) if isinstance(v, (dict,list)) else v} for k,v in audit.items()]))
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        headers = list(rows[0]) if rows else ["empty"]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
        ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78"); cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            letter = col[0].column_letter
            ws.column_dimensions[letter].width = min(max(max(len(str(c.value or "")) for c in col) + 2, 12), 58)
        if rows:
            tab = Table(displayName=f"T_{title}", ref=ws.dimensions)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tab)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    args = parser.parse_args()
    output_dir = args.output_dir; output_dir.mkdir(parents=True, exist_ok=True)
    db_path = output_dir / "synthetic_chart_of_accounts_tree_v1.sqlite"
    nodes, relations = build_nodes()
    create_db(db_path, nodes, relations)
    audit = validate(db_path)
    if audit["status"] != "PASS":
        raise RuntimeError(canonical(audit))
    db_sha = sha256_file(db_path)
    write_design(output_dir / "会计科目树V1设计.md", audit, db_sha)
    write_excel(output_dir / "会计科目树V1_人工审查.xlsx", db_path, audit)
    (output_dir / "会计科目树构建审计.json").write_text(canonical({"asset_sha256": db_sha, "audit": audit}) + "\n", encoding="utf-8")
    print(canonical({"output_dir": str(output_dir), "asset_sha256": db_sha, "audit": audit}))


if __name__ == "__main__":
    main()
