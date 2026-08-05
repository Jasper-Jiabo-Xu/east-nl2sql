#!/usr/bin/env python3
"""Build query-ready V3 code libraries: one logical code set, one SQLite table.

The previous V2 generic-row library is intentionally not read here.  This builder
preserves raw source rows but creates a physical table per Attachment-4 code table.
For multi-level code sheets (notably 行业类型), it completes the ancestor columns
before inserting rows, while retaining `raw_values_json` for audit.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from vnext.constraint_assets_v1.compile_tasks import ATT3, parse_att3


ROOT = Path(__file__).resolve().parents[2]
WORKING = ROOT / "kb/working/20260727_000_goal_mode_v1"
CODEBOOK = ROOT / "data/raw/20260626_east_materials/east材料/规范附件4：金融资产管理公司监管数据标准化规范业务代码表.xlsx"
REVIEW_BOOK = WORKING / "构建过程层/01_附件3主数据与附件2技术规则解析/附件3附件2解析审阅报告.xlsx"
DEFAULT_OUTPUT = WORKING / "构建过程层/04_字段多字段与对象明细状态提取-V2/code_value_libraries/20260804_一表一码表_层级补全_v3"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def sha_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def source_ref(source: str, sheet: str, row: int, final_column: str) -> str:
    return f"{source}:{sheet}!A{row}:{final_column}{row}"


def safe_table_name(code_table_id: str, sheet_name: str) -> str:
    pinyin_fallback = re.sub(r"[^a-zA-Z0-9]+", "_", sheet_name).strip("_").lower() or "code_set"
    return f"code_{code_table_id.lower()}_{pinyin_fallback}"[:60]


def unique_headers(headers: list[str]) -> list[str]:
    result, used = [], set()
    for ordinal, header in enumerate(headers, start=1):
        base = re.sub(r"\s+", "", header) or f"原始列{ordinal}"
        candidate, suffix = base, 2
        while candidate in used:
            candidate = f"{base}_{suffix}"
            suffix += 1
        result.append(candidate)
        used.add(candidate)
    return result


def merged_values(sheet) -> list[list[str]]:
    """Return a value grid with merged-cell values expanded to all covered cells."""
    rows = [[norm(cell.value) for cell in row] for row in sheet.iter_rows()]
    for merged in sheet.merged_cells.ranges:
        value = norm(sheet.cell(merged.min_row, merged.min_col).value)
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                rows[row - 1][column - 1] = value
    return rows


def layout(rows: list[list[str]], *, has_hierarchy_subheader: bool = False) -> tuple[list[int], list[int], list[str], list[int]]:
    """Identify header rows, effective headers, and hierarchy columns."""
    if len(rows) < 2:
        raise RuntimeError("码表sheet不足两行")
    # Attachment-4 regular tables use title row + header row. Only 行业类型 has
    # a second header row whose cells name the four hierarchy levels.  Do not
    # infer this from first data-row text: CD000013's first value description
    # contains “代码”, which previously made it a false subheader.
    second = rows[1]
    third = rows[2] if len(rows) > 2 else []
    third_is_subheader = has_hierarchy_subheader
    header_rows = [2, 3] if third_is_subheader else [2]
    effective = []
    for index in range(max(len(second), len(third))):
        upper = second[index] if index < len(second) else ""
        lower = third[index] if third_is_subheader and index < len(third) else ""
        effective.append(lower or upper)
    source_columns = [index for index, header in enumerate(effective) if header and header != "返回目录"]
    headers = unique_headers([effective[index] for index in source_columns])
    hierarchy_columns = [source_columns.index(index) for index, value in enumerate(second) if value == "代码取值" and index in source_columns] if third_is_subheader else []
    return header_rows, source_columns, headers, hierarchy_columns


def fill_hierarchy(raw: list[str], hierarchy_columns: list[int], carry: list[str]) -> tuple[list[str], list[str], int | None, str | None, str | None]:
    """Complete parent codes.  A new value clears all lower-level carry values."""
    filled = list(raw)
    deepest = None
    for position, column in enumerate(hierarchy_columns):
        value = raw[column] if column < len(raw) else ""
        if value:
            carry[position] = value
            for later in range(position + 1, len(carry)):
                carry[later] = ""
            deepest = position
        filled[column] = carry[position]
    if deepest is None:
        deepest = max((index for index, value in enumerate(carry) if value), default=-1)
    if deepest < 0:
        return filled, carry, None, None, None
    path = [carry[index] for index in range(deepest + 1) if carry[index]]
    parent = carry[deepest - 1] if deepest > 0 else None
    return filled, carry, deepest + 1, parent, "/".join(path)


def create_metadata(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE source_document (
          source_id TEXT PRIMARY KEY,
          absolute_path TEXT NOT NULL,
          sha256 TEXT NOT NULL,
          imported_at TEXT NOT NULL
        );
        CREATE TABLE code_table_registry (
          code_table_id TEXT PRIMARY KEY,
          sqlite_table_name TEXT NOT NULL UNIQUE,
          sheet_name TEXT NOT NULL UNIQUE,
          reference_standard_raw TEXT NOT NULL,
          directory_source_ref TEXT NOT NULL UNIQUE,
          hierarchy_column_names_json TEXT NOT NULL,
          header_names_json TEXT NOT NULL,
          row_count INTEGER NOT NULL,
          source_id TEXT NOT NULL REFERENCES source_document(source_id)
        );
        CREATE TABLE external_standard (
          external_standard_id TEXT PRIMARY KEY,
          east_citation TEXT NOT NULL,
          standard_name TEXT NOT NULL,
          verification_status TEXT NOT NULL,
          source_summary TEXT NOT NULL,
          authoritative_link TEXT,
          use_decision TEXT NOT NULL,
          source_ref TEXT NOT NULL UNIQUE
        );
        CREATE TABLE data_element_code_table_binding (
          binding_id TEXT PRIMARY KEY,
          external_standard_id TEXT NOT NULL REFERENCES external_standard(external_standard_id),
          east_citation TEXT NOT NULL,
          data_element_code TEXT NOT NULL,
          data_element_name TEXT NOT NULL,
          att3_location TEXT NOT NULL,
          att3_quote TEXT NOT NULL,
          code_table_id TEXT REFERENCES code_table_registry(code_table_id),
          binding_status TEXT NOT NULL CHECK(binding_status IN ('BOUND','NO_ATT4_TABLE')),
          source_ref TEXT NOT NULL UNIQUE,
          UNIQUE(external_standard_id, data_element_code, code_table_id)
        );
        CREATE INDEX idx_v3_binding_element ON data_element_code_table_binding(data_element_code);
        """
    )


def create_code_table(conn: sqlite3.Connection, table_name: str, headers: list[str], hierarchy_columns: list[int], extra_columns: list[str] | None = None) -> None:
    fields = [
        '"source_row" INTEGER NOT NULL UNIQUE',
        '"source_ref" TEXT NOT NULL UNIQUE',
        '"hierarchy_level" INTEGER',
        '"parent_code" TEXT',
        '"code_path" TEXT',
        '"raw_values_json" TEXT NOT NULL',
    ] + [f"{quote(header)} TEXT" for header in headers] + [f"{quote(column)} TEXT" for column in (extra_columns or [])]
    conn.execute(f"CREATE TABLE {quote(table_name)} ({', '.join(fields)})")
    for column in hierarchy_columns:
        conn.execute(f"CREATE INDEX {quote(table_name + '_h' + str(column))} ON {quote(table_name)} ({quote(headers[column])})")


ADMIN_DIVISION_COLUMNS = [
    "行政区划层级", "名称前导空格数",
    "省级行政区划代码", "省级行政区划名称",
    "地市级行政区划代码", "地市级行政区划名称",
    "县区级行政区划代码", "县区级行政区划名称",
]


def administrative_division_hierarchy(
    code: str,
    name_with_indent: str,
    carry: dict[str, tuple[str, str] | None],
) -> tuple[dict[str, tuple[str, str] | None], int, str, str | None, str, list[str]]:
    """Use the source B-column leading spaces to complete province/city/county.

    Attachment 4 uses no leading space for provincial-level rows, one space for
    prefecture/city rows and three spaces for county/district rows.  Municipalities
    directly under the central government therefore have a province row followed
    directly by level-3 county/district rows; their city fields remain null.
    """
    leading_spaces = len(name_with_indent) - len(name_with_indent.lstrip(" "))
    name = name_with_indent.strip()
    if leading_spaces == 0:
        level, label = 1, "PROVINCE"
        carry = {"province": (code, name), "prefecture": None, "county": None}
    elif leading_spaces == 1:
        if carry["province"] is None:
            raise RuntimeError(f"行政区划地市行缺少省级父级: {code} {name}")
        level, label = 2, "PREFECTURE"
        carry = {**carry, "prefecture": (code, name), "county": None}
    elif leading_spaces >= 3:
        if carry["province"] is None:
            raise RuntimeError(f"行政区划县区行缺少省级父级: {code} {name}")
        level, label = 3, "COUNTY"
        carry = {**carry, "county": (code, name)}
    else:
        raise RuntimeError(f"行政区划名称缩进无法识别（预期0、1或至少3空格）: {code} {name_with_indent!r}")
    parent = carry["prefecture"] if level == 3 and carry["prefecture"] else (carry["province"] if level > 1 else None)
    path = "/".join(item[0] for item in (carry["province"], carry["prefecture"], carry["county"]) if item)
    extra = [
        label, str(leading_spaces),
        *(value for key in ("province", "prefecture", "county") for pair in (carry[key],) for value in ((pair[0], pair[1]) if pair else (None, None))),
    ]
    return carry, level, label, parent[0] if parent else None, path, extra


def import_reference_library(path: Path) -> dict[str, Any]:
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys=ON")
    create_metadata(conn)
    imported_at = now()
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_AMC_CODE", str(CODEBOOK.resolve()), sha_file(CODEBOOK), imported_at))
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_REVIEW", str(REVIEW_BOOK.resolve()), sha_file(REVIEW_BOOK), imported_at))

    review = load_workbook(REVIEW_BOOK, read_only=True, data_only=True)
    for row_no, row in enumerate(review["外部标准候选"].iter_rows(min_row=2, values_only=True), start=2):
        values = [norm(value) for value in row[:7]]
        if values[0]:
            conn.execute("INSERT INTO external_standard VALUES (?,?,?,?,?,?,?,?)", (*values, source_ref("SRC_REVIEW", "外部标准候选", row_no, "G")))

    codebook = load_workbook(CODEBOOK, read_only=False, data_only=True)
    directory = codebook["目录"]
    directory_map: dict[str, tuple[str, str, int]] = {}
    for row_no, row in enumerate(directory.iter_rows(min_row=3, values_only=True), start=3):
        code, name, standard = norm(row[0]), norm(row[1]), norm(row[2])
        if code and name:
            directory_map[name] = (code, standard, row_no)

    table_count = value_rows = 0
    directory_unmatched_sheets: list[str] = []
    for sheet_name in codebook.sheetnames:
        if sheet_name == "目录":
            continue
        if sheet_name in directory_map:
            code_table_id, standard, directory_row = directory_map[sheet_name]
        else:
            code_table_id, standard, directory_row = f"SHEET_{hashlib.sha256(sheet_name.encode()).hexdigest()[:12]}", "目录未精确对应（人工审核）", 0
            directory_unmatched_sheets.append(sheet_name)
        sheet = codebook[sheet_name]
        rows = merged_values(sheet)
        header_rows, source_columns, headers, hierarchy_columns = layout(rows, has_hierarchy_subheader=code_table_id == "CD000012")
        table_name = safe_table_name(code_table_id, sheet_name)
        is_admin_division = code_table_id == "CD000003"
        extra_columns = ADMIN_DIVISION_COLUMNS if is_admin_division else []
        create_code_table(conn, table_name, headers, hierarchy_columns, extra_columns)
        carry = [""] * len(hierarchy_columns)
        admin_carry: dict[str, tuple[str, str] | None] = {"province": None, "prefecture": None, "county": None}
        admin_name_index = headers.index("代码名称") if is_admin_division else -1
        admin_code_index = headers.index("代码取值") if is_admin_division else -1
        inserted = 0
        for row_no, source_values in enumerate(rows[max(header_rows):], start=max(header_rows) + 1):
            raw = [source_values[index] if index < len(source_values) else "" for index in source_columns]
            if not any(raw) or raw[0] == "返回目录":
                continue
            # A row must contain either a code position or a name.  This excludes
            # residual heading/return rows but preserves sparse hierarchy rows.
            if not any(raw[index] for index in hierarchy_columns) and not any("名称" not in value and value for value in raw):
                continue
            extra_values: list[str | None] = []
            if is_admin_division:
                # `merged_values` normalizes strings, so retrieve source B directly
                # to retain the literal leading spaces that encode hierarchy.
                raw_name = "" if sheet.cell(row_no, source_columns[admin_name_index] + 1).value is None else str(sheet.cell(row_no, source_columns[admin_name_index] + 1).value)
                raw_for_audit = list(raw)
                raw_for_audit[admin_name_index] = raw_name
                admin_carry, level, _, parent, path, extra_values = administrative_division_hierarchy(raw[admin_code_index], raw_name, admin_carry)
                filled = list(raw)
                filled[admin_name_index] = raw_name.strip()
            else:
                raw_for_audit = raw
                filled, carry, level, parent, path = fill_hierarchy(raw, hierarchy_columns, carry) if hierarchy_columns else (raw, carry, None, None, None)
            final_column = chr(64 + min(len(headers), 26))
            values = [row_no, source_ref("SRC_AMC_CODE", sheet_name, row_no, final_column), level, parent, path, json.dumps(raw_for_audit, ensure_ascii=False)] + filled + extra_values
            placeholders = ",".join("?" for _ in values)
            columns = ["source_row", "source_ref", "hierarchy_level", "parent_code", "code_path", "raw_values_json"] + headers + extra_columns
            conn.execute(f"INSERT INTO {quote(table_name)} ({','.join(quote(column) for column in columns)}) VALUES ({placeholders})", values)
            inserted += 1
        registry_ref = source_ref("SRC_AMC_CODE", "目录", directory_row, "C") if directory_row else f"SRC_AMC_CODE:{sheet_name}!A1"
        conn.execute("INSERT INTO code_table_registry VALUES (?,?,?,?,?,?,?,?,?)", (code_table_id, table_name, sheet_name, standard, registry_ref, json.dumps([headers[index] for index in hierarchy_columns], ensure_ascii=False), json.dumps(headers, ensure_ascii=False), inserted, "SRC_AMC_CODE"))
        table_count += 1
        value_rows += inserted

    for row_no, row in enumerate(review["国标与数据元绑定"].iter_rows(min_row=2, values_only=True), start=2):
        external_id, citation, element_code, element_name, location, att3_quote, code_table_id = [norm(value) for value in row[:7]]
        if not external_id:
            continue
        status = "BOUND" if code_table_id else "NO_ATT4_TABLE"
        binding_id = "BND_" + hashlib.sha256(f"{external_id}|{element_code}|{code_table_id}".encode()).hexdigest()[:24]
        conn.execute("INSERT INTO data_element_code_table_binding VALUES (?,?,?,?,?,?,?,?,?,?)", (binding_id, external_id, citation, element_code, element_name, location, att3_quote, code_table_id or None, status, source_ref("SRC_REVIEW", "国标与数据元绑定", row_no, "G")))
    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
    no_duplicate_registry = conn.execute("SELECT COUNT(*) FROM (SELECT sqlite_table_name, COUNT(*) n FROM code_table_registry GROUP BY sqlite_table_name HAVING n > 1)").fetchone()[0]
    industry_table = conn.execute("SELECT sqlite_table_name FROM code_table_registry WHERE code_table_id='CD000012'").fetchone()[0]
    industry_incomplete = conn.execute(f"SELECT COUNT(*) FROM {quote(industry_table)} WHERE {quote('行业门类')}='' OR {quote('行业门类')} IS NULL").fetchone()[0]
    counts = {"code_table_count": table_count, "code_value_row_count": value_rows, "binding_count": conn.execute("SELECT COUNT(*) FROM data_element_code_table_binding").fetchone()[0], "industry_incomplete_category_rows": industry_incomplete}
    conn.commit()
    conn.close()
    return {"integrity_check": integrity, "duplicate_physical_table_names": no_duplicate_registry, "directory_unmatched_sheets": directory_unmatched_sheets, **counts}


def import_local_source_library(path: Path) -> dict[str, Any]:
    """Create V3 local-code registry. One physical table will be created per set by V3 importer."""
    conn = sqlite3.connect(path)
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(
        """
        CREATE TABLE source_document (source_id TEXT PRIMARY KEY, absolute_path TEXT NOT NULL, sha256 TEXT NOT NULL, imported_at TEXT NOT NULL);
        CREATE TABLE local_code_set_registry (
          local_code_set_id TEXT PRIMARY KEY,
          sqlite_table_name TEXT NOT NULL UNIQUE,
          owner_kind TEXT NOT NULL,
          owner_id TEXT NOT NULL,
          owner_name TEXT NOT NULL,
          source_ref TEXT NOT NULL UNIQUE,
          raw_text TEXT NOT NULL,
          extraction_status TEXT NOT NULL CHECK(extraction_status IN ('PENDING_AGENT1_V3','EXTRACTED','MANUAL_REVIEW')),
          UNIQUE(owner_kind, owner_id)
        );
        """
    )
    conn.execute("INSERT INTO source_document VALUES (?,?,?,?)", ("SRC_ATT3", str(ATT3.resolve()), sha_file(ATT3), now()))
    elements, _, _ = parse_att3()
    for element in elements.values():
        match = re.search(r"!A(\d+):", element["source_refs"][0])
        if not match:
            raise RuntimeError(f"无法定位数据元来源行: {element['code']}")
        row = int(match.group(1))
        set_id = f"LCS_DE_{element['code']}"
        table_name = f"local_de_{element['code']}"
        conn.execute("INSERT INTO local_code_set_registry VALUES (?,?,?,?,?,?,?,?)", (set_id, table_name, "DATA_ELEMENT_DESCRIPTION", element["code"], element["name"], f"SRC_ATT3:数据元说明!E{row}", element["description"], "PENDING_AGENT1_V3"))
    workbook = load_workbook(ATT3, read_only=True, data_only=True)
    active_table = ""
    remarks_count = 0
    for row_no, row in enumerate(workbook["数据结构"].iter_rows(values_only=True), start=1):
        values = list(row[:16])
        if re.fullmatch(r"[A-Z][A-Z0-9]*", norm(values[2] if len(values) > 2 else "")):
            active_table = norm(values[2])
        field_code, field_name, remarks = norm(values[6] if len(values) > 6 else ""), norm(values[5] if len(values) > 5 else ""), norm(values[12] if len(values) > 12 else "")
        if not active_table or not re.fullmatch(r"[A-Z][A-Z0-9]*", field_code) or not remarks:
            continue
        field_id = f"FLD_{active_table}_{field_code}"
        conn.execute("INSERT INTO local_code_set_registry VALUES (?,?,?,?,?,?,?,?)", (f"LCS_{field_id}", f"local_{active_table.lower()}_{field_code.lower()}", "FIELD_REMARK", field_id, field_name, f"SRC_ATT3:数据结构!M{row_no}", remarks, "PENDING_AGENT1_V3"))
        remarks_count += 1
    integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
    source_count = conn.execute("SELECT COUNT(*) FROM local_code_set_registry").fetchone()[0]
    conn.commit()
    conn.close()
    return {"integrity_check": integrity, "local_code_set_registry_count": source_count, "data_element_sources": len(elements), "field_remark_sources": remarks_count, "physical_local_value_tables": 0, "next_step": "V3 Agent1 validated result creates each local_code_set_registry.sqlite_table_name physical table; no V2 draft value is imported."}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    output = args.output_dir.resolve()
    output.mkdir(parents=True, exist_ok=True)
    reference = output / "code_reference_library_v3.sqlite"
    local = output / "local_code_value_library_v3.sqlite"
    if reference.exists() or local.exists():
        raise SystemExit("目标已存在；请指定新的输出目录，避免覆盖审查资产")
    audit = {"created_at": now(), "publication_status": "review_only_not_released", "supersedes": "20260804_附件4码表与本地码值骨架（通用行仓库，不作为后续输入）", "reference_library": import_reference_library(reference), "local_value_library": import_local_source_library(local)}
    (output / "code_value_library_v3_audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
