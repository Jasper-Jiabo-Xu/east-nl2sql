# -*- coding: utf-8 -*-
"""EAST V5 审阅工作簿生成器（确定性，仅脱敏 synthetic fixture）。

生成 Sol 冻结的四类工作表：
  - 审阅总览
  - 30候选明细（30 槽位 = 5 题 × 6 baseline）
  - Q1..Q5（每题 1 个完整问题 + 6 路 baseline SQL）
  - 运行与校验证据

硬约束实现：
  - 问题/SQL 单元格一律 set_explicit_value(data_type='s') + number_format='@' 强制纯文本；
  - 冻结窗格、自动筛选、自动换行；
  - 人工结论/状态使用受限下拉值（DataValidation list）；
  - 字节级确定性输出（固定 zip 时间戳 + 固定文档属性）。

用法：
    python3 generate_review_workbook.py [输出路径]
默认输出 ./review_workbook_template.xlsx
"""

import os
import re
import sys
import hashlib
import json
import zipfile
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

import review_workbook_spec as spec

FIXED_TS = datetime(2026, 1, 1, 0, 0, 0, tzinfo=timezone.utc)


# --------------------------------------------------------------------------- #
# synthetic fixture（确定性）
# --------------------------------------------------------------------------- #
SYNTH_QUESTIONS = [
    "[合成示例] 查询客户表中状态为活跃的客户名称与客户编号。",
    "[合成示例] 查询账户表中余额大于 1000 的账户列表。",
    "[合成示例] 按机构维度统计商户数量并输出分组结果。",
    "[合成示例] 关联商户表与机构表，查询每个商户所属的机构名称。",
    "[合成示例] 查询最近 30 天内有交易记录的商户及其最近一笔交易时间。",
]

SYNTH_QA_IDS = ["SYNTH_QA_%03d" % i for i in range(1, spec.NUM_QUESTIONS + 1)]


def synth_original_sql(qi: int, bi: int) -> str:
    return (
        "-- %s baseline=%s\n"
        "SELECT synth_col_%d FROM synth_table_%d WHERE synth_id = %d;"
        % (spec.SYNTH_MARK, spec.BASELINES[bi - 1], qi, qi, bi)
    )


def synth_normalized_sql(qi: int, bi: int) -> str:
    return "select synth_col_%d from synth_table_%d where synth_id = %d;" % (qi, qi, bi)


def synth_hash(label: str) -> str:
    return hashlib.sha256(("synthetic-" + label).encode("utf-8")).hexdigest()


# 唯一失败槽位：Q5 × AutoLink（演示失败码/失败摘要路径）
FAIL_SLOT = (5, 5)  # (qi=5 -> index 4, bi=5 -> index 5)


def slot_record(qi: int, bi: int):
    """返回 (qa_id, baseline, 原始SQL, 规范化SQL, 状态, 失败码, SQL哈希, 备注)。"""
    qa_id = SYNTH_QA_IDS[qi - 1]
    baseline = spec.BASELINES[bi - 1]
    if (qi, bi) == FAIL_SLOT:
        return (qa_id, baseline, "", "", "失败", "SYNTH_FAIL_001", "", "模拟失败槽位")
    raw = synth_original_sql(qi, bi)
    norm = synth_normalized_sql(qi, bi)
    return (qa_id, baseline, raw, norm, "成功", "", hashlib.sha256(raw.encode("utf-8")).hexdigest(), "")


def fixture_digest():
    """synthetic fixture 的 canonical 内容哈希（生成器与校验器共用）。"""
    canon = {
        "spec_baselines": spec.BASELINES,
        "questions": SYNTH_QUESTIONS,
        "qa_ids": SYNTH_QA_IDS,
        "slots": [
            {
                "qa_id": SYNTH_QA_IDS[qi - 1],
                "baseline": spec.BASELINES[bi - 1],
                "raw_sql": slot_record(qi, bi)[2],
                "norm_sql": slot_record(qi, bi)[3],
                "status": slot_record(qi, bi)[4],
                "failure_code": slot_record(qi, bi)[5],
            }
            for qi in range(1, spec.NUM_QUESTIONS + 1)
            for bi in range(1, spec.NUM_BASELINES + 1)
        ],
    }
    return hashlib.sha256(
        json.dumps(canon, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


# --------------------------------------------------------------------------- #
# 单元格写入
# --------------------------------------------------------------------------- #
def write_text(ws, row, col, text):
    """强制纯文本：data_type='s' + 文本格式 + 自动换行。防公式注入。"""
    c = ws.cell(row=row, column=col)
    c.value = str(text)
    c.data_type = "s"  # 即使首字符为 =/+/-/@ 也强制文本，不落公式
    c.number_format = "@"
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


def write_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=str(text))
    c.font = Font(bold=True)
    c.alignment = Alignment(vertical="top")
    return c


def style_header(ws, row, ncols):
    fill = PatternFill("solid", fgColor="DDEBF7")
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")


def add_list_validation(ws, values, cell_range):
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(values),
        allow_blank=True,
        showDropDown=False,  # False => 显示下拉箭头（openpyxl 语义反转）
    )
    dv.error = "请从受限值中选择：%s" % "/".join(values)
    dv.errorTitle = "非法取值"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(cell_range)


# --------------------------------------------------------------------------- #
# 各表
# --------------------------------------------------------------------------- #
def build_overview(ws, fixture_hex, s0_hash, adapter_hash, run_hash):
    ws.title = spec.SHEET_OVERVIEW
    headers = ["项目", "值", "说明"]
    for i, h in enumerate(headers, 1):
        write_label(ws, 1, i, h)
    style_header(ws, 1, len(headers))

    success = spec.TOTAL_SLOTS - 1
    rows = [
        ("模型合同", "synthetic-model-contract-v1.0", "%s 模型合同占位" % spec.SYNTH_MARK),
        ("S0 哈希", s0_hash, "%s synthetic-s0" % spec.SYNTH_MARK),
        ("adapter 哈希", adapter_hash, "%s synthetic-adapter" % spec.SYNTH_MARK),
        ("run 哈希", run_hash, "%s synthetic-run" % spec.SYNTH_MARK),
        ("workbook 哈希", "__PENDING__", "由校验器在 manifest 中回填，避免自引用"),
        ("30 槽位完成度", "%d/%d 成功，1 失败" % (success, spec.TOTAL_SLOTS), spec.SYNTH_MARK),
        ("失败摘要", "1 失败：Q5 × AutoLink（SYNTH_FAIL_001）", spec.SYNTH_MARK),
    ]
    r = 2
    for item, val, note in rows:
        write_label(ws, r, 1, item)
        write_text(ws, r, 2, val)
        write_text(ws, r, 3, note)
        r += 1
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 44
    ws.freeze_panes = "A2"


def build_detail(ws):
    ws.title = spec.SHEET_DETAIL
    for i, h in enumerate(spec.DETAIL_HEADERS, 1):
        write_label(ws, 1, i, h)
    style_header(ws, 1, len(spec.DETAIL_HEADERS))

    r = 2
    for qi in range(1, spec.NUM_QUESTIONS + 1):
        for bi in range(1, spec.NUM_BASELINES + 1):
            qa_id, baseline, raw, norm, status, fail_code, sql_hash, note = slot_record(qi, bi)
            row_vals = {
                1: qi,                  # 题序
                2: qa_id,               # qa_id
                3: SYNTH_QUESTIONS[qi - 1],  # 完整问题
                4: baseline,            # baseline
                5: raw,                 # 原始SQL
                6: norm,                # 规范化SQL
                7: status,              # 状态
                8: fail_code,           # 失败码
                9: sql_hash,            # SQL哈希
                10: "synth-model-v1.0",
                11: "synth-adapter-v1.0",
                12: "synth-upstream-v1.0",
                13: 1,                  # attempt
                14: 1,                  # call
                15: (qi * 100 + bi * 7),    # token（确定性）
                16: (qi * 10 + bi * 3),     # latency ms（确定性）
                17: "trace://synth/%s/%s" % (qa_id, baseline),
                18: "待审阅" if (qi, bi) == FAIL_SLOT else "通过",  # 人工结论
                19: note,
            }
            for col, val in row_vals.items():
                if col in spec.DETAIL_TEXT_COLS:
                    write_text(ws, r, col, val)
                else:
                    c = ws.cell(row=r, column=col, value=val)
                    c.alignment = Alignment(vertical="top", wrap_text=(col == 19))
            r += 1

    last_row = 1 + spec.TOTAL_SLOTS
    # 下拉：状态 G、人工结论 R
    add_list_validation(ws, spec.STATUS_VALUES, "G2:G%d" % last_row)
    add_list_validation(ws, spec.REVIEW_CONCLUSION_VALUES, "R2:R%d" % last_row)

    widths = [6, 14, 40, 22, 46, 40, 8, 16, 64, 16, 18, 18, 8, 6, 8, 8, 26, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"  # 冻结行 1 + 列 A/B（题序、qa_id）
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(spec.DETAIL_HEADERS)), last_row)


def build_q_sheet(ws, qi):
    ws.title = spec.q_sheet_name(qi)
    qa_id = SYNTH_QA_IDS[qi - 1]
    question = SYNTH_QUESTIONS[qi - 1]

    write_label(ws, 1, 1, "qa_id")
    write_text(ws, 1, 2, qa_id)
    write_label(ws, 2, 1, "题序")
    ws.cell(row=2, column=2, value=qi)
    write_label(ws, 3, 1, "完整问题")
    qc = write_text(ws, 3, 2, question)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)

    # 表头
    for i, h in enumerate(spec.Q_HEADERS, 1):
        write_label(ws, spec.Q_HEADER_ROW, i, h)
    style_header(ws, spec.Q_HEADER_ROW, len(spec.Q_HEADERS))

    for bi in range(1, spec.NUM_BASELINES + 1):
        r = spec.Q_FIRST_DATA_ROW + bi - 1
        _qa, baseline, raw, norm, status, fail_code, sql_hash, _note = slot_record(qi, bi)
        write_text(ws, r, 1, baseline)
        write_text(ws, r, 2, raw)
        write_text(ws, r, 3, norm)
        c = ws.cell(row=r, column=4, value=status)
        c.alignment = Alignment(vertical="top")
        ws.cell(row=r, column=5, value=fail_code)
        ws.cell(row=r, column=6, value=sql_hash)

    add_list_validation(
        ws, spec.STATUS_VALUES,
        "D%d:D%d" % (spec.Q_FIRST_DATA_ROW, spec.Q_LAST_DATA_ROW),
    )

    for col, w in zip("ABCDEF", [22, 46, 40, 8, 16, 64]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A%d" % (spec.Q_HEADER_ROW + 1)
    ws.auto_filter.ref = "A%d:F%d" % (spec.Q_HEADER_ROW, spec.Q_LAST_DATA_ROW)


def build_evidence(ws):
    ws.title = spec.SHEET_EVIDENCE
    for i, h in enumerate(["项目", "值", "说明"], 1):
        write_label(ws, 1, i, h)
    style_header(ws, 1, 3)

    rows = [
        ("输入版本", "synthetic-fixture-v1.0", spec.SYNTH_MARK),
        ("精确 model ID", "synthetic-model-id（未解析占位）", "%s 不调用真实模型" % spec.SYNTH_MARK),
        ("provider 合同", "synthetic-provider-contract-v1.0", spec.SYNTH_MARK),
        ("时间", "2026-08-27T00:00:00Z", "%s 固定时间戳" % spec.SYNTH_MARK),
        ("唯一性校验", "30/30 唯一（qa_id×baseline）", "待校验器复核"),
        ("计数校验", "5 题 × 6 baseline = 30 槽位", "待校验器复核"),
        ("泄漏扫描", "0 泄漏（仅 synthetic fixture）", "待校验器复核"),
        ("工作簿校验", "通过（结构/公式注入/下拉/冻结/筛选）", "待校验器复核"),
    ]
    r = 2
    for item, val, note in rows:
        write_label(ws, r, 1, item)
        write_text(ws, r, 2, val)
        write_text(ws, r, 3, note)
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A2"


# --------------------------------------------------------------------------- #
# 确定性保存
# --------------------------------------------------------------------------- #
def save_deterministic(wb, path):
    tmp = path + ".tmp"
    wb.save(tmp)
    fixed_modified = "2026-01-01T00:00:00Z"
    with zipfile.ZipFile(tmp, "r") as zin:
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "docProps/core.xml":
                    # openpyxl 在保存时会把 modified 覆写为当前时间，这里归一化为固定值
                    text = data.decode("utf-8")
                    text = re.sub(
                        r"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                        r"\g<1>%s\g<2>" % fixed_modified,
                        text,
                    )
                    data = text.encode("utf-8")
                item.date_time = (1980, 1, 1, 0, 0, 0)
                item.compress_type = zipfile.ZIP_DEFLATED
                item.create_system = 3
                item.external_attr = 0o100644 << 16
                zout.writestr(item, data)
    os.remove(tmp)


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "review_workbook_template.xlsx"
    s0_hash = synth_hash("s0")
    adapter_hash = synth_hash("adapter")
    run_hash = synth_hash("run")
    fixture_hex = fixture_digest()

    wb = Workbook()
    ws = wb.active
    build_overview(ws, fixture_hex, s0_hash, adapter_hash, run_hash)
    build_detail(wb.create_sheet())
    for qi in range(1, spec.NUM_QUESTIONS + 1):
        build_q_sheet(wb.create_sheet(), qi)
    build_evidence(wb.create_sheet())

    # 固定文档属性，保证 docProps 确定性
    wb.properties.creator = "EAST-V5-review-workbook"
    wb.properties.title = "EAST V5 审阅工作簿（synthetic fixture）"
    wb.properties.created = FIXED_TS
    wb.properties.modified = FIXED_TS

    save_deterministic(wb, out)

    # 输出 meta（供证据）
    meta = {
        "output": out,
        "fixture_sha256": fixture_hex,
        "s0_sha256": s0_hash,
        "adapter_sha256": adapter_hash,
        "run_sha256": run_hash,
        "workbook_sha256": hashlib.sha256(open(out, "rb").read()).hexdigest(),
        "sheets": spec.SHEET_ORDER,
        "total_slots": spec.TOTAL_SLOTS,
    }
    print(json.dumps(meta, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
