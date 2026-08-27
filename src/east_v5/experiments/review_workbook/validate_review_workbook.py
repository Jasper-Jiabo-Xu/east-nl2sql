# -*- coding: utf-8 -*-
"""EAST V5 审阅工作簿校验器（确定性，只读，不修改工作簿）。

校验 Sol 冻结结构 + 硬约束：
  1. 工作表存在与顺序；
  2. 30候选明细 19 列头精确匹配；
  3. 30 槽位 = 5 qa_id × 6 baseline 且 (qa_id,baseline) 唯一；
  4. 六路 baseline 精确匹配；
  5. 问题/SQL 单元格强制纯文本（data_type != 'f'，number_format='@'），公式注入扫描；
  6. 冻结窗格 / 自动筛选 / 自动换行；
  7. 审阅结论与状态受限下拉值；
  8. 状态↔失败码、SQL 哈希一致性；
  9. 泄漏扫描（密钥模式 + synthetic 标记核验）。

输出 manifest JSON 到 <输出目录>/review_workbook.manifest.json，并在 stdout 打印摘要。
任一硬检查失败则进程退出码非 0。

用法：
    python3 validate_review_workbook.py <workbook.xlsx> [输出目录]
"""

import os
import sys
import json
import hashlib
import re

from openpyxl import load_workbook

import review_workbook_spec as spec

# 常见密钥/凭据模式（泄漏扫描，只做本地脱敏结论，不外发）
_SECRET_PATTERNS = [
    re.compile(r"sk-[A-Za-z0-9]{8,}"),
    re.compile(r"AKIA[0-9A-Z]{16}"),
    re.compile(r"ghp_[A-Za-z0-9]{20,}"),
    re.compile(r"gho_[A-Za-z0-9]{20,}"),
    re.compile(r"github_pat_[A-Za-z0-9_]{20,}"),
    re.compile(r"xox[bap]-[A-Za-z0-9-]{10,}"),
    re.compile(r"-----BEGIN [A-Z ]*PRIVATE KEY-----"),
    re.compile(r"eyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}"),
    re.compile(r"[a-z]+://[^/]+:[^/@]+@"),
]


def check(name, ok, detail=""):
    return {"name": name, "pass": bool(ok), "detail": detail}


def cell_is_formula(c):
    return getattr(c, "data_type", None) == "f"


def text_cells_ok(ws, coords):
    """校验问题/SQL 单元格：非公式 + 文本格式。返回 (ok, 违规列表)。"""
    bad = []
    for (r, c) in coords:
        cell = ws.cell(row=r, column=c)
        if cell_is_formula(cell):
            bad.append("formula@%s%d" % (cell.coordinate, cell.row))
        elif cell.value is not None and cell.number_format != "@":
            bad.append("numfmt@%s" % cell.coordinate)
    return (len(bad) == 0, bad)


def collect_dvs(ws):
    """返回 [(formula1, [range_str, ...]), ...]"""
    out = []
    try:
        for dv in ws.data_validations.dataValidation:
            f1 = dv.formula1 or ""
            ranges = []
            sqref = dv.sqref
            if sqref is None:
                continue
            # MultiCellRange 可迭代为 CellRange
            try:
                ranges = [str(r) for r in sqref.ranges]
            except AttributeError:
                ranges = [str(sqref)]
            out.append((f1, ranges))
    except Exception:
        pass
    return out


def dv_covers(ws, expected_values, target_range):
    """是否存在一个下拉 DV，其取值集合 == expected_values 且覆盖 target_range。"""
    target = target_range.split(":")[0]  # 仅校验起点单元格
    for f1, ranges in collect_dvs(ws):
        vals = [v.strip() for v in f1.strip('"').split(",")]
        if set(vals) == set(expected_values):
            for rng in ranges:
                # rng 形如 "$G$2:$G$31" 或 "G2:G31"
                norm = rng.replace("$", "")
                if norm == target_range or norm.startswith(target + ":"):
                    return True
    return False


def detail_checks(ws):
    checks = []
    # 列头
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(spec.DETAIL_HEADERS) + 1)]
    checks.append(check(
        "30候选明细 列头精确匹配",
        headers == spec.DETAIL_HEADERS,
        "got=%r" % headers,
    ))

    rows = ws.max_row
    data_rows = rows - 1
    checks.append(check(
        "30候选明细 数据行数=30",
        data_rows == spec.TOTAL_SLOTS,
        "data_rows=%d" % data_rows,
    ))

    slots = []
    qa_ids = []
    for r in range(2, rows + 1):
        qa_id = ws.cell(row=r, column=2).value
        baseline = ws.cell(row=r, column=4).value
        slots.append((qa_id, baseline, r))
        qa_ids.append(qa_id)

    unique = set((s[0], s[1]) for s in slots)
    checks.append(check(
        "30 槽位 qa_id×baseline 唯一",
        len(unique) == spec.TOTAL_SLOTS and len(unique) == len(slots),
        "unique_pairs=%d/%d" % (len(unique), len(slots)),
    ))

    got_baselines = sorted(set(s[1] for s in slots))
    checks.append(check(
        "六路 baseline 精确匹配",
        got_baselines == sorted(spec.BASELINES),
        "got=%r" % got_baselines,
    ))

    got_qa = sorted(set(s[0] for s in slots))
    checks.append(check(
        "5 个唯一 qa_id",
        len(got_qa) == spec.NUM_QUESTIONS and all(q for q in got_qa),
        "qa_ids=%r" % got_qa,
    ))

    # 纯文本 + 公式注入
    text_coords = [(r, c) for r in range(2, rows + 1) for c in spec.DETAIL_TEXT_COLS]
    ok, bad = text_cells_ok(ws, text_coords)
    checks.append(check("30候选明细 问题/SQL 纯文本", ok, "violations=%d %r" % (len(bad), bad[:5])))

    # 公式注入扫描：全表是否存在任何公式单元格
    formula_cells = []
    for r in range(1, rows + 1):
        for c in range(1, len(spec.DETAIL_HEADERS) + 1):
            if cell_is_formula(ws.cell(row=r, column=c)):
                formula_cells.append(ws.cell(row=r, column=c).coordinate)
    checks.append(check("30候选明细 无公式单元格", len(formula_cells) == 0, "formulas=%r" % formula_cells[:5]))

    # 状态↔失败码一致性
    inconsistent = []
    for r in range(2, rows + 1):
        status = ws.cell(row=r, column=spec.DETAIL_STATUS_COL).value
        fail = ws.cell(row=r, column=8).value
        if status == "失败" and not fail:
            inconsistent.append("row%d: 失败但失败码空" % r)
        if status == "成功" and fail:
            inconsistent.append("row%d: 成功但失败码非空" % r)
    checks.append(check("状态↔失败码一致性", len(inconsistent) == 0, "inconsistent=%r" % inconsistent[:5]))

    # SQL 哈希一致性（成功槽位）
    hash_bad = []
    for r in range(2, rows + 1):
        status = ws.cell(row=r, column=spec.DETAIL_STATUS_COL).value
        raw = ws.cell(row=r, column=5).value
        h = ws.cell(row=r, column=9).value
        if status == "成功":
            expect = hashlib.sha256((raw or "").encode("utf-8")).hexdigest()
            if h != expect:
                hash_bad.append("row%d" % r)
    checks.append(check("SQL 哈希一致性（成功槽位）", len(hash_bad) == 0, "bad=%r" % hash_bad[:5]))

    # 格式化
    checks.append(check("30候选明细 冻结窗格", ws.freeze_panes is not None, "freeze=%r" % ws.freeze_panes))
    checks.append(check("30候选明细 自动筛选", ws.auto_filter.ref is not None, "ref=%r" % ws.auto_filter.ref))
    wrap_ok = all(
        ws.cell(row=r, column=c).alignment.wrap_text
        for r in range(2, rows + 1) for c in spec.DETAIL_TEXT_COLS
    )
    checks.append(check("30候选明细 自动换行", wrap_ok, ""))

    checks.append(check(
        "人工结论受限下拉（R 列）",
        dv_covers(ws, spec.REVIEW_CONCLUSION_VALUES, "R2:R%d" % rows),
        "expected=%r" % spec.REVIEW_CONCLUSION_VALUES,
    ))
    checks.append(check(
        "状态受限下拉（G 列）",
        dv_covers(ws, spec.STATUS_VALUES, "G2:G%d" % rows),
        "expected=%r" % spec.STATUS_VALUES,
    ))

    return checks, got_qa


def q_sheet_checks(ws, qi):
    checks = []
    qa_id = ws.cell(row=1, column=2).value
    checks.append(check("Q%d qa_id 存在" % qi, bool(qa_id), "qa_id=%r" % qa_id))

    # 6 路 baseline
    baselines = [ws.cell(row=r, column=1).value for r in range(spec.Q_FIRST_DATA_ROW, spec.Q_LAST_DATA_ROW + 1)]
    checks.append(check(
        "Q%d 六路 baseline 完整" % qi,
        sorted(baselines) == sorted(spec.BASELINES),
        "got=%r" % baselines,
    ))

    # 纯文本
    coords = [(r, c) for r in range(spec.Q_FIRST_DATA_ROW, spec.Q_LAST_DATA_ROW + 1) for c in spec.Q_TEXT_COLS]
    ok, bad = text_cells_ok(ws, coords)
    checks.append(check("Q%d SQL 纯文本" % qi, ok, "violations=%d %r" % (len(bad), bad[:5])))

    checks.append(check("Q%d 冻结窗格" % qi, ws.freeze_panes is not None, "freeze=%r" % ws.freeze_panes))
    checks.append(check("Q%d 自动筛选" % qi, ws.auto_filter.ref is not None, "ref=%r" % ws.auto_filter.ref))
    wrap_ok = all(
        ws.cell(row=r, column=c).alignment.wrap_text
        for r in range(spec.Q_FIRST_DATA_ROW, spec.Q_LAST_DATA_ROW + 1) for c in spec.Q_TEXT_COLS
    )
    checks.append(check("Q%d 自动换行" % qi, wrap_ok, ""))
    checks.append(check(
        "Q%d 状态下拉" % qi,
        dv_covers(ws, spec.STATUS_VALUES, "D%d:D%d" % (spec.Q_FIRST_DATA_ROW, spec.Q_LAST_DATA_ROW)),
        "expected=%r" % spec.STATUS_VALUES,
    ))
    return checks, qa_id


def leak_scan(wb):
    """扫描全部单元格字符串，返回违规坐标列表（密钥模式）。"""
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    for p in _SECRET_PATTERNS:
                        if p.search(v):
                            hits.append("%s!%s" % (ws.title, cell.coordinate))
                            break
    return hits


def synth_mark_check(wb):
    """确认所有完整问题单元格带 synthetic 标记（证明仅脱敏 fixture）。"""
    missing = []
    ws = wb[spec.SHEET_DETAIL]
    for r in range(2, ws.max_row + 1):
        q = ws.cell(row=r, column=3).value
        if q and spec.SYNTH_MARK not in str(q):
            missing.append("detail!C%d" % r)
    for qi in range(1, spec.NUM_QUESTIONS + 1):
        wsq = wb[spec.q_sheet_name(qi)]
        q = wsq.cell(row=3, column=2).value
        if q and spec.SYNTH_MARK not in str(q):
            missing.append("Q%d!B3" % qi)
    return missing


def main():
    if len(sys.argv) < 2:
        print("用法: python3 validate_review_workbook.py <workbook.xlsx> [输出目录]")
        sys.exit(2)
    path = sys.argv[1]
    outdir = sys.argv[2] if len(sys.argv) > 2 else "."

    wb = load_workbook(path, data_only=False)
    checks = []

    # 工作表顺序
    got_order = [ws.title for ws in wb.worksheets]
    checks.append(check(
        "工作表存在与顺序",
        got_order == spec.SHEET_ORDER,
        "got=%r" % got_order,
    ))

    got_qa = []
    if spec.SHEET_DETAIL in wb.sheetnames:
        dchecks, got_qa = detail_checks(wb[spec.SHEET_DETAIL])
        checks.extend(dchecks)

    q_qa = []
    for qi in range(1, spec.NUM_QUESTIONS + 1):
        name = spec.q_sheet_name(qi)
        if name in wb.sheetnames:
            qchecks, qa_id = q_sheet_checks(wb[name], qi)
            checks.extend(qchecks)
            q_qa.append(qa_id)

    # 明细 qa_id 与 Q 表 qa_id 一致性
    if got_qa and q_qa:
        checks.append(check(
            "明细 qa_id 与 Q1..Q5 一致",
            sorted(set(got_qa)) == sorted(set(q_qa)),
            "detail=%r q=%r" % (sorted(set(got_qa)), sorted(set(q_qa))),
        ))

    # 泄漏扫描
    hits = leak_scan(wb)
    checks.append(check("密钥/凭据泄漏扫描（0 命中）", len(hits) == 0, "hits=%r" % hits[:5]))

    # synthetic 标记
    missing_mark = synth_mark_check(wb)
    checks.append(check(
        "synthetic 标记核验（所有完整问题带 %s）" % spec.SYNTH_MARK,
        len(missing_mark) == 0,
        "missing=%r" % missing_mark[:5],
    ))

    # 工作簿哈希
    wb_bytes = open(path, "rb").read()
    workbook_sha256 = hashlib.sha256(wb_bytes).hexdigest()

    passed = [c for c in checks if c["pass"]]
    failed = [c for c in checks if not c["pass"]]

    manifest = {
        "schema": "east-v5-review-workbook-manifest-v1",
        "workbook": os.path.basename(path),
        "workbook_sha256": workbook_sha256,
        "total_checks": len(checks),
        "passed": len(passed),
        "failed": len(failed),
        "result": "PASS" if len(failed) == 0 else "FAIL",
        "checks": checks,
        "leak_scan_hits": hits,
    }
    os.makedirs(outdir, exist_ok=True)
    out_path = os.path.join(outdir, "review_workbook.manifest.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(json.dumps({
        "workbook": os.path.basename(path),
        "workbook_sha256": workbook_sha256,
        "result": manifest["result"],
        "passed": len(passed),
        "failed": len(failed),
    }, ensure_ascii=False, indent=2))
    for c in failed:
        print("FAIL: %s — %s" % (c["name"], c["detail"]))
    print("manifest:", out_path)

    sys.exit(0 if len(failed) == 0 else 1)


if __name__ == "__main__":
    main()
